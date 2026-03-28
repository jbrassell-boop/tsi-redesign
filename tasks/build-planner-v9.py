import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

hdr_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='143B87')
n = Font(name='Arial', size=10)
b = Font(name='Arial', bold=True, size=10)
sec = Font(name='Arial', bold=True, size=12, color='143B87')
thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
thick = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
money = '$#,##0;($#,##0);"-"'
pct = '0.0%'
yellow = PatternFill('solid', fgColor='FFFF00')
green = PatternFill('solid', fgColor='C6EFCE')
light = PatternFill('solid', fgColor='D6E4F0')
orange = PatternFill('solid', fgColor='FCE4D6')
navy = PatternFill('solid', fgColor='143B87')
nf = Font(name='Arial', bold=True, size=11, color='FFFFFF')
alt = PatternFill('solid', fgColor='F2F2F2')
red = Font(name='Arial', bold=True, size=16, color='FF0000')

def H(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True); cell.border = thin

def S(ws, r, c, val=None, fmt=None, bold=False):
    cell = ws.cell(r, c)
    if val is not None: cell.value = val
    cell.font = b if bold else n; cell.border = thin
    if fmt: cell.number_format = fmt
    return cell

def NR(ws, r, cols):
    for c in range(1, cols + 1):
        ws.cell(r, c).font = nf; ws.cell(r, c).fill = navy; ws.cell(r, c).border = thin

# ============================================================
# SHEET 1: Revenue Planner
# ============================================================
ws = wb.active
ws.title = 'Revenue Planner'
ws.sheet_properties.tabColor = '143B87'

ws['A1'] = 'TSI Revenue Planner'
ws['A1'].font = Font(name='Arial', bold=True, size=20, color='143B87')
ws.merge_cells('A1:G1')

# Goal
S(ws, 3, 1, 'Revenue Goal', bold=True).font = Font(name='Arial', bold=True, size=14)
c = S(ws, 3, 2, 20000000, money)
c.fill = yellow; c.border = thick; c.font = Font(name='Arial', bold=True, size=18, color='FF0000')

# Current Roster
r = 5
ws.cell(r, 1, 'WHAT WE HAVE NOW').font = sec
r += 1
for c, h in enumerate(['Rep', 'Type', 'Annual Revenue'], 1): ws.cell(r, c, h)
H(ws, r, 3)
r += 1

current = [
    ('Brian Kenney', 'Inside', 1455000),
    ('Debbie Hightower', 'Inside', 1627084),
    ('Bernie DeLacy', 'Outside', 1750205),
    ('Brandi Cook', 'Inside', 1815826),
    ('Seamus Glavin', 'Inside', 1139552),
    ('Jim Rygiel', 'Outside', 442609),
    ('Timothy Reilly', 'Inside', 487810),
    ('Donald Haynes', 'Inside', 125999),
    ('Ryan George', 'Inside', 238269),
    ('BSC Programs', 'Program', 2198651),
]

rf = r
for name, rtype, rev in current:
    S(ws, r, 1, name, bold=True)
    S(ws, r, 2, rtype)
    c = S(ws, r, 3, rev, money); c.fill = yellow
    fill = green if rtype == 'Outside' else orange if rtype == 'Program' else light
    ws.cell(r, 1).fill = fill; ws.cell(r, 2).fill = fill
    r += 1
rl = r - 1

S(ws, r, 1, 'CURRENT TOTAL', bold=True)
S(ws, r, 3, f'=SUM(C{rf}:C{rl})', money)
NR(ws, r, 3)
ct = r; r += 1

S(ws, r, 1, 'GAP TO GOAL', bold=True)
S(ws, r, 3, f'=B3-C{ct}', money).font = red
ws.cell(r, 1).font = Font(name='Arial', bold=True, size=14, color='FF0000')
r += 2

# Add Reps
ws.cell(r, 1, 'ADD REPS TO CLOSE THE GAP').font = sec
r += 1
ws.cell(r, 1, 'Pick: "Inherited" (taking over a book), "Cold" (net new only), or "Territory" (outside territory rep). Revenue fills in.').font = Font(name='Arial', italic=True, size=9, color='666666')
ws.merge_cells(f'A{r}:G{r}')
r += 1

for c, h in enumerate(['New Rep', 'Scenario', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'], 1):
    ws.cell(r, c, h)
H(ws, r, 7)
hh = r; r += 1

# Benchmarks in reference rows (55-57)
# Inherited (2015+ avg): total revenue with book
inh_bench = [275826, 304129, 399683, 411662, 469603]
# Cold start (2015+ avg net new only)
cold_bench = [43126, 70273, 106364, 87704, 85735]
# Outside (using all-years avg since 2015+ only has 1)
out_bench = [348374, 404261, 491083, 814695, 1455687]

for i, v in enumerate(inh_bench): S(ws, 55, i + 1, v, money)  # Inherited
for i, v in enumerate(cold_bench): S(ws, 56, i + 1, v, money)  # Cold
for i, v in enumerate(out_bench): S(ws, 57, i + 1, v, money)   # Outside
S(ws, 55, 6, 'Inherited (avg, 23 reps, 2015+)')
S(ws, 56, 6, 'Cold Start / Net New Only (avg, 23 reps, 2015+)')
S(ws, 57, 6, 'Territory (outside rep avg, 8 reps, all years)')

hf = r
for i in range(10):
    S(ws, r, 1, f'New Rep {i+1}')
    c = S(ws, r, 2, 'Inherited' if i < 3 else 'Cold' if i < 5 else 'Territory' if i < 7 else '')
    c.fill = yellow
    for yr in range(1, 6):
        col = yr + 2
        ref = get_column_letter(yr)
        S(ws, r, col, f'=IF(B{r}="Inherited",{ref}55,IF(B{r}="Cold",{ref}56,IF(B{r}="Territory",{ref}57,0)))', money)
    if i % 2 == 0:
        for ci in range(1, 8):
            if ws.cell(r, ci).fill != yellow: ws.cell(r, ci).fill = alt
    r += 1
hl = r - 1

S(ws, r, 1, 'NEW HIRES TOTAL', bold=True)
for col in range(3, 8):
    cl = get_column_letter(col)
    S(ws, r, col, f'=SUM({cl}{hf}:{cl}{hl})', money)
NR(ws, r, 7)
ht = r; r += 2

# The Answer
ws.cell(r, 1, 'THE ANSWER').font = Font(name='Arial', bold=True, size=16, color='143B87')
r += 1
for c, h in enumerate(['', '', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'], 1): ws.cell(r, c, h)
H(ws, r, 7); r += 1

S(ws, r, 1, 'Current Roster', bold=True)
for col in range(3, 8): S(ws, r, col, f'=C{ct}', money)
for ci in range(1, 8): ws.cell(r, ci).fill = light
r += 1

S(ws, r, 1, 'New Hire Revenue', bold=True)
for col in range(3, 8): S(ws, r, col, f'={get_column_letter(col)}{ht}', money)
for ci in range(1, 8): ws.cell(r, ci).fill = green
r += 1

S(ws, r, 1, 'TOTAL REVENUE', bold=True)
for col in range(3, 8): S(ws, r, col, f'={get_column_letter(col)}{r-2}+{get_column_letter(col)}{r-1}', money)
NR(ws, r, 7)
pt = r; r += 1

S(ws, r, 1, 'Goal', bold=True)
for col in range(3, 8): S(ws, r, col, '=$B$3', money)
r += 1

S(ws, r, 1, 'OVER / (SHORT)', bold=True)
ws.cell(r, 1).font = red
for col in range(3, 8):
    S(ws, r, col, f'={get_column_letter(col)}{pt}-{get_column_letter(col)}{r-1}', money).font = Font(name='Arial', bold=True, size=14, color='FF0000')

ws.column_dimensions['A'].width = 28; ws.column_dimensions['B'].width = 14
for c in range(3, 8): ws.column_dimensions[get_column_letter(c)].width = 16

# ============================================================
# SHEET 2: What to Expect
# ============================================================
ws2 = wb.create_sheet('What to Expect')
ws2.sheet_properties.tabColor = '70AD47'

ws2['A1'] = 'What to Expect from a New Rep: Years 1-5'
ws2['A1'].font = Font(name='Arial', bold=True, size=18, color='143B87')
ws2.merge_cells('A1:F1')
ws2['A2'] = 'Based on 23 inside reps hired 2015-2025 | Median removes outliers | "How We Got This" shows the source'
ws2['A2'].font = Font(name='Arial', italic=True, size=10, color='666666')
ws2.merge_cells('A2:F2')

# Scenario 1: Rep inherits a book
r = 4
ws2.cell(r, 1, 'SCENARIO 1: REP INHERITS A BOOK').font = Font(name='Arial', bold=True, size=14, color='143B87')
ws2.merge_cells(f'A{r}:F{r}')
r += 1
ws2.cell(r, 1, 'Rep takes over existing clients/territory. Like Seamus, Hightower, Guillermo. Most TSI hires.').font = Font(name='Arial', italic=True, size=9, color='666666')
ws2.merge_cells(f'A{r}:F{r}')
r += 1

for c, h in enumerate(['Metric', 'Year 1', 'Year 2', 'Year 3', 'How We Got This'], 1): ws2.cell(r, c, h)
H(ws2, r, 5); r += 1

inh_metrics = [
    ('Total Revenue', '$276K', '$304K', '$400K', 'Avg of 23 inside reps (2015+), FFS + contract'),
    ('  Inherited Revenue', '$192K', '$209K', '$223K', 'Revenue from clients that existed before this rep'),
    ('  Net New Revenue', '$43K', '$70K', '$106K', 'Revenue from clients this rep brought in'),
    ('  Contract Revenue', '$41K', '$24K', '$71K', 'Contract billing attributed to this rep'),
    ('Net New % of Total', '16%', '23%', '27%', 'Net new grows each year as rep builds own book'),
    ('Total Clients', '25', '25', '27', 'Clients actively sending work'),
    ('  Net New Clients', '7', '9', '9', 'Clients the rep personally acquired'),
    ('  Inherited Clients', '18', '16', '18', 'Clients handed to the rep'),
    ('Top Performers Yr1', '', '', '', 'Seamus $1.5M, Hightower $667K, Rygiel $436K, Guillermo $434K'),
    ('Median Yr1', '$216K', '', '', '50th percentile — half above, half below'),
]
for i, (metric, y1, y2, y3, source) in enumerate(inh_metrics):
    S(ws2, r, 1, metric, bold=not metric.startswith(' '))
    S(ws2, r, 2, y1); S(ws2, r, 3, y2); S(ws2, r, 4, y3)
    S(ws2, r, 5, source).font = Font(name='Arial', italic=True, size=9, color='666666')
    if i % 2 == 0:
        for ci in range(1, 6): ws2.cell(r, ci).fill = green
    r += 1

# Scenario 2: Cold start
r += 1
ws2.cell(r, 1, 'SCENARIO 2: COLD START (no inherited book)').font = Font(name='Arial', bold=True, size=14, color='143B87')
ws2.merge_cells(f'A{r}:F{r}')
r += 1
ws2.cell(r, 1, 'Rep placed in new territory with no existing clients. Like sending someone to Texas/California cold.').font = Font(name='Arial', italic=True, size=9, color='666666')
ws2.merge_cells(f'A{r}:F{r}')
r += 1

for c, h in enumerate(['Metric', 'Year 1', 'Year 2', 'Year 3', 'How We Got This'], 1): ws2.cell(r, c, h)
H(ws2, r, 5); r += 1

cold_metrics = [
    ('Total Revenue', '$43K', '$70K', '$106K', 'Net new only — no inherited clients to start'),
    ('  = All Net New', '$43K', '$70K', '$106K', 'Everything comes from rep\'s own hunting'),
    ('  Contract Revenue', '$0', '$0', '$0', 'Contracts take time — typically Year 3+ if at all'),
    ('Net New Clients', '7', '9', '9', 'Same client acquisition rate, just no base'),
    ('Avg Revenue/New Client', '$6K', '$8K', '$12K', 'Revenue per client grows as relationships deepen'),
    ('Breakeven (est)', 'Month 10-14', '', '', 'At $65-80K comp, need ~$70K revenue to break even'),
    ('Top Net New Yr1', '', '', '', 'Rygiel $263K, Hightower $160K, Reilly $158K, Guillermo $109K'),
    ('Realistic Yr1 Range', '$15-50K', '', '', 'Most cold-start reps: $15-50K Year 1, then ramp Year 2-3'),
]
for i, (metric, y1, y2, y3, source) in enumerate(cold_metrics):
    S(ws2, r, 1, metric, bold=not metric.startswith(' '))
    S(ws2, r, 2, y1); S(ws2, r, 3, y2); S(ws2, r, 4, y3)
    S(ws2, r, 5, source).font = Font(name='Arial', italic=True, size=9, color='666666')
    if i % 2 == 0:
        for ci in range(1, 6): ws2.cell(r, ci).fill = PatternFill('solid', fgColor='FFC7CE')
    r += 1

# Side by side
r += 1
ws2.cell(r, 1, 'SIDE BY SIDE: INHERITED vs COLD START').font = Font(name='Arial', bold=True, size=14, color='FF0000')
ws2.merge_cells(f'A{r}:F{r}')
r += 1
for c, h in enumerate(['', 'Inherited Book', 'Cold Start', 'Difference'], 1): ws2.cell(r, c, h)
H(ws2, r, 4); r += 1

sbs = [
    ('Year 1 Revenue', '$276K', '$43K', '$233K gap'),
    ('Year 2 Revenue', '$304K', '$70K', '$234K gap'),
    ('Year 3 Revenue', '$400K', '$106K', '$294K gap'),
    ('Year 1 Clients', '25', '7', '18 fewer'),
    ('Time to $200K/yr', 'Day 1', 'Year 3+', '2+ year head start'),
    ('Contract Revenue (Yr3)', '$71K', '$0', 'Inherited reps land contracts faster'),
    ('Risk Level', 'Lower', 'Higher', 'Cold start = longer to ROI'),
]
for i, (metric, inh, cold, diff) in enumerate(sbs):
    S(ws2, r, 1, metric, bold=True); S(ws2, r, 2, inh); S(ws2, r, 3, cold); S(ws2, r, 4, diff)
    if i % 2 == 0:
        for ci in range(1, 5): ws2.cell(r, ci).fill = alt
    r += 1

# Bottom line
r += 1
ws2.cell(r, 1, 'BOTTOM LINE FOR DENIS').font = Font(name='Arial', bold=True, size=14, color='FF0000')
ws2.merge_cells(f'A{r}:F{r}')
r += 1
lines = [
    'A rep taking over a book = $276K Year 1. A rep starting cold = $43K Year 1. That\'s a 6x difference.',
    'Both reps add the same number of net new clients (~7-9/yr). The gap is entirely the inherited base.',
    'Cold start reps break even around Month 10-14. Inherited reps are profitable from Day 1.',
    'Contract revenue appears by Year 3 for inherited reps ($71K avg). Cold start reps don\'t get there.',
    'The smart play: give new hires a book wherever possible. Even a small inherited base accelerates everything.',
    'If you MUST go cold (new territory): budget for 12-18 months of investment before meaningful returns.',
]
for i, line in enumerate(lines):
    S(ws2, r, 1, f'  {i+1}. {line}')
    ws2.merge_cells(f'A{r}:F{r}')
    r += 1

# Data sources
r += 1
ws2.cell(r, 1, 'DATA SOURCES').font = sec; r += 1
sources = [
    '23 inside reps hired 2015-2025, each with 20+ invoices and 6+ months active.',
    'FFS revenue from tblInvoice (200K+ rows). Contract billing from tblContractDepartmentInvoiceSchedule ($35M).',
    'Net new = client\'s first-ever invoice was under this rep. Inherited = client existed before.',
    'Contract revenue attributed to the contract\'s own lSalesRepKey (not the client\'s rep).',
    'Current roster revenue averaged 2024-2025 to smooth spikes (e.g. Surgical Solutions departure).',
    'Reps: Seamus $1.5M, Hightower $667K, Rygiel $436K, Guillermo $434K, O\'Connell $397K, Reilly $378K...',
]
for s in sources:
    S(ws2, r, 1, f'  - {s}'); ws2.merge_cells(f'A{r}:F{r}'); r += 1

for c in range(1, 6): ws2.column_dimensions[get_column_letter(c)].width = 16
ws2.column_dimensions['A'].width = 30; ws2.column_dimensions['E'].width = 55

# ============================================================
# SHEET 3: Every Rep Year 1 (proof)
# ============================================================
ws3 = wb.create_sheet('Rep Year 1 Detail')
ws3.sheet_properties.tabColor = '4472C4'

ws3['A1'] = 'Every Inside Rep Year 1 (2015-2025) — The Raw Data'
ws3['A1'].font = Font(name='Arial', bold=True, size=16, color='143B87')
ws3.merge_cells('A1:G1')
ws3['A2'] = '23 reps sorted by total revenue. This is exactly what each rep generated in their first 12 months.'
ws3['A2'].font = Font(name='Arial', italic=True, size=10, color='666666')
ws3.merge_cells('A2:G2')

for c, h in enumerate(['Rep', 'Total Revenue', 'Inherited', 'Net New', 'Contract', 'Clients', 'Net New Clients'], 1):
    ws3.cell(4, c, h)
H(ws3, 4, 7)

reps_yr1 = [
    ('Seamus Glavin', 1525321, 924420, 23611, 577290, 65, 3),
    ('Debbie Hightower', 666698, 506427, 160272, 0, 70, 32),
    ('Jim Rygiel', 435991, 172697, 263293, 0, 10, 9),
    ('Fausto Guillermo', 433644, 158365, 109016, 166263, 31, 7),
    ('Andrew O\'Connell', 397103, 387065, 10038, 0, 29, 3),
    ('Timothy Reilly', 377670, 177425, 158044, 42200, 40, 17),
    ('Giuseppe Gambino', 345955, 343113, 2842, 0, 35, 5),
    ('Meghan Weaver', 334125, 295461, 38664, 0, 39, 10),
    ('Chris Bower', 282019, 247781, 34238, 0, 29, 8),
    ('Brandi Cook', 280516, 120558, 1288, 158670, 15, 2),
    ('Paul Crossin', 262220, 254714, 7506, 0, 25, 3),
    ('Tom Kane', 215860, 175913, 39947, 0, 20, 4),
    ('Gabrielle Santorio', 191218, 158873, 32345, 0, 23, 7),
    ('Michael Strommen', 171950, 167395, 4556, 0, 30, 4),
    ('Charles Kenney', 111874, 99449, 12425, 0, 25, 6),
    ('Bill Hoffman', 74472, 74472, 0, 0, 14, 0),
    ('John T Sargent', 58930, 40285, 18645, 0, 21, 10),
    ('Carla Goodwin', 51899, 39694, 12205, 0, 12, 6),
    ('Brandi Cook (2nd key)', 50833, 30619, 20214, 0, 15, 11),
    ('Phil T Cerami', 40333, 11628, 28705, 0, 16, 10),
    ('Meghan Wright', 30793, 16748, 14045, 0, 9, 5),
    ('Tom Kane (2nd key)', 4575, 4575, 0, 0, 1, 0),
    ('Scott Holloway', 0, 0, 0, 0, 4, 0),
]

for i, (name, total, inh, nn, con, cl, nnc) in enumerate(reps_yr1):
    row = 5 + i
    S(ws3, row, 1, name, bold=True)
    S(ws3, row, 2, total, money)
    S(ws3, row, 3, inh, money)
    S(ws3, row, 4, nn, money)
    S(ws3, row, 5, con, money)
    S(ws3, row, 6, cl)
    S(ws3, row, 7, nnc)
    if i % 2 == 0:
        for ci in range(1, 8): ws3.cell(row, ci).fill = alt

# Averages
row = 5 + len(reps_yr1) + 1
S(ws3, row, 1, 'AVERAGE', bold=True)
for col in range(2, 8):
    cl = get_column_letter(col)
    S(ws3, row, col, f'=AVERAGE({cl}5:{cl}{4+len(reps_yr1)})', money if col < 6 else '0')
NR(ws3, row, 7)
row += 1
S(ws3, row, 1, 'MEDIAN', bold=True)
for col in range(2, 8):
    cl = get_column_letter(col)
    S(ws3, row, col, f'=MEDIAN({cl}5:{cl}{4+len(reps_yr1)})', money if col < 6 else '0')
for ci in range(1, 8): ws3.cell(row, ci).fill = green

for c in range(1, 8): ws3.column_dimensions[get_column_letter(c)].width = 16
ws3.column_dimensions['A'].width = 24

# Save
out = r'C:\tmp\tsi-sales-rep-planner-v9.xlsx'
wb.save(out)
print(f'Saved to {out}')
print('Sheets:', [s.title for s in wb.worksheets])
