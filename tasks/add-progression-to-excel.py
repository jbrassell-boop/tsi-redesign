import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(r'C:\tmp\tsi-sales-rep-benchmarks.xlsx')

hdr_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='143B87')
sub_fill = PatternFill('solid', fgColor='4472C4')
normal_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', bold=True, size=10)
section_font = Font(name='Arial', bold=True, size=12, color='143B87')
title_font = Font(name='Arial', bold=True, size=14, color='143B87')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
money_fmt = '$#,##0;($#,##0);"-"'
pct_fmt = '0.0%'
green_fill = PatternFill('solid', fgColor='C6EFCE')
light_fill = PatternFill('solid', fgColor='D6E4F0')

def style_header(ws, row, max_col, fill=None):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = hdr_font
        cell.fill = fill or hdr_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def sc(ws, r, c, val=None, fmt=None, bold=False):
    cell = ws.cell(row=r, column=c)
    if val is not None: cell.value = val
    cell.font = bold_font if bold else normal_font
    cell.border = thin_border
    if fmt: cell.number_format = fmt
    return cell

# ============================================================
# NEW SHEET: Year 1-5 Progression Dashboard
# ============================================================
ws = wb.create_sheet('Yr 1-5 Progression', 0)
ws.sheet_properties.tabColor = '143B87'

ws['A1'] = 'Sales Rep Year 1-5 Progression Dashboard'
ws['A1'].font = Font(name='Arial', bold=True, size=16, color='143B87')
ws.merge_cells('A1:I1')
ws['A2'] = 'What to expect when placing a rep in a territory | 109 reps, 25 years of invoice data'
ws['A2'].font = Font(name='Arial', italic=True, size=11, color='666666')
ws.merge_cells('A2:I2')

# Section 1: Revenue Progression
r = 4
ws.cell(r, 1, 'REVENUE PROGRESSION').font = section_font
r += 1
headers = ['', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5']
for c, h in enumerate(headers, 1): ws.cell(r, c, h)
style_header(ws, r, 6)

rows = [
    ('Total Revenue (avg)', [231914, 274479, 340148, 377677, 492651]),
    ('  Net New Revenue', [44085, 74028, 146008, 176736, 257577]),
    ('  Inherited Revenue', [187829, 200451, 194140, 200941, 235074]),
    ('Net New % of Total', [0.19, 0.27, 0.43, 0.47, 0.52]),
    ('# Reps in Sample', [101, 77, 46, 33, 22]),
]
for i, (label, vals) in enumerate(rows):
    row = r + 1 + i
    sc(ws, row, 1, label, bold='Total' in label or 'Net New %' in label)
    if label == 'Net New % of Total':
        for c, v in enumerate(vals, 2): sc(ws, row, c, v, pct_fmt)
    elif label == '# Reps in Sample':
        for c, v in enumerate(vals, 2): sc(ws, row, c, v, '#,##0')
    else:
        for c, v in enumerate(vals, 2): sc(ws, row, c, v, money_fmt)
    if i == 0:
        for c in range(1, 7): ws.cell(row, c).fill = light_fill
    if label == 'Net New % of Total':
        for c in range(1, 7): ws.cell(row, c).fill = green_fill

# Section 2: FFS vs Contract
r += 9
ws.cell(r, 1, 'FFS vs CONTRACT REVENUE').font = section_font
r += 1
headers2 = ['', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5']
for c, h in enumerate(headers2, 1): ws.cell(r, c, h)
style_header(ws, r, 6)

ffs_rows = [
    ('Fee-for-Service (avg)', [188021, 237889, 299592, 312096, 439992]),
    ('Contract Revenue (avg)', [9532, 8594, 28965, 32581, 36167]),
    ('Contract % of Total', [0.05, 0.03, 0.09, 0.09, 0.08]),
    ('# Reps in Sample', [118, 84, 49, 37, 24]),
]
for i, (label, vals) in enumerate(ffs_rows):
    row = r + 1 + i
    sc(ws, row, 1, label, bold='FFS' in label or 'Contract %' in label)
    if 'Contract %' in label:
        for c, v in enumerate(vals, 2): sc(ws, row, c, v, pct_fmt)
    elif '# Reps' in label:
        for c, v in enumerate(vals, 2): sc(ws, row, c, v, '#,##0')
    else:
        for c, v in enumerate(vals, 2): sc(ws, row, c, v, money_fmt)

ws.cell(r + 6, 1, 'Key Insight: Contract revenue is minimal in years 1-5. Reps live on FFS/T&M early.').font = Font(name='Arial', italic=True, size=10, color='FF0000')
ws.merge_cells(f'A{r+6}:F{r+6}')

# Section 3: Unique Clients
r += 9
ws.cell(r, 1, 'UNIQUE CLIENT COUNT').font = section_font
r += 1
for c, h in enumerate(['', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5', 'Year 6', 'Year 7', 'Year 8'], 1):
    ws.cell(r, c, h)
style_header(ws, r, 9)

client_counts = [33, 37, 49, 36, 44, 53, 53, 56]
row = r + 1
sc(ws, row, 1, 'Avg Unique Clients', bold=True)
for c, v in enumerate(client_counts, 2): sc(ws, row, c, v, '#,##0')

# Section 4: Dominant Instrument Type
r += 5
ws.cell(r, 1, 'INSTRUMENT MIX (most reps)').font = section_font
r += 1
for c, h in enumerate(['Category', '% of Business', 'Notes'], 1): ws.cell(r, c, h)
style_header(ws, r, 3)
inst_data = [
    ('Flexible Endoscopes', '45-75%', 'Dominant category for most outside reps'),
    ('Rigid Scopes', '15-30%', 'Second largest, higher avg ticket'),
    ('Cameras/Instruments', '5-15%', 'Accessories and camera heads'),
    ('Contract Coverage', '3-15%', 'Capitated/shared risk repairs'),
]
for i, (cat, pct, note) in enumerate(inst_data):
    row = r + 1 + i
    sc(ws, row, 1, cat, bold=True)
    sc(ws, row, 2, pct)
    sc(ws, row, 3, note)

# Section 5: The Pitch to Denis
r += 8
ws.cell(r, 1, 'BOTTOM LINE FOR NEW TERRITORY HIRE').font = Font(name='Arial', bold=True, size=14, color='FF0000')
ws.merge_cells(f'A{r}:F{r}')
r += 1

pitch = [
    ('Cold Territory (no book)', '$44K', '$74K', '$146K', '$177K', '$258K'),
    ('With Inherited Book', '$232K', '$274K', '$340K', '$378K', '$493K'),
    ('Crossover Point', '', '', '', '', 'Year 5: Net new > inherited'),
    ('Breakeven (est @ $80K comp)', 'Month 8-10', '', '', '', ''),
    ('First Contract (typical)', '', 'Year 2-3', '', '', '$30-90K/yr recurring'),
    ('Client Ramp', '19 clients', '19 clients', '22 clients', '23 clients', '27 clients'),
]
for c, h in enumerate(['Scenario', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'], 1):
    ws.cell(r, c, h)
style_header(ws, r, 6)
for i, (label, *vals) in enumerate(pitch):
    row = r + 1 + i
    sc(ws, row, 1, label, bold=True)
    for c, v in enumerate(vals, 2): sc(ws, row, c, v)
    if i == 0:
        for c in range(1, 7): ws.cell(row, c).fill = PatternFill('solid', fgColor='FFC7CE')
    elif i == 1:
        for c in range(1, 7): ws.cell(row, c).fill = green_fill

# Column widths
for c in range(1, 10):
    ws.column_dimensions[get_column_letter(c)].width = 18
ws.column_dimensions['A'].width = 28

# ============================================================
# Update Rep Directory with years active & instrument type
# ============================================================
ws_dir = wb['Rep Directory']
# Add instrument type column header
ws_dir.cell(4, 8, 'Dominant Instrument')
ws_dir.cell(4, 8).font = hdr_font
ws_dir.cell(4, 8).fill = hdr_fill
ws_dir.cell(4, 8).alignment = Alignment(horizontal='center', wrap_text=True)
ws_dir.cell(4, 8).border = thin_border

ws_dir.cell(4, 9, 'Years Active')
ws_dir.cell(4, 9).font = hdr_font
ws_dir.cell(4, 9).fill = hdr_fill
ws_dir.cell(4, 9).alignment = Alignment(horizontal='center', wrap_text=True)
ws_dir.cell(4, 9).border = thin_border

# Instrument types and years for key reps
rep_extra = {
    'Brian Kenney': ('F(38%), R(32%), I(22%)', '15yr'),
    'Debbie Hightower': ('F(41%), R(32%), I(15%)', '27yr'),
    'John Luedke': ('F(53%), R(21%), C(13%)', '24yr'),
    'Gina Ellis': ('F(48%), R(25%), I(15%)', '7yr'),
    'Lindsey Davis': ('F(54%), R(28%), C(10%)', '5yr'),
    'Ryan George': ('F(71%), O(19%), R(9%)', '9yr'),
    'Maria Mercanti': ('F(60%), R(22%), C(7%)', '7yr'),
    'Jim Hueston': ('F(53%), R(19%), C(16%)', '7yr'),
    'Ryan Gades': ('F(58%), R(27%), C(13%)', '12yr'),
    'Fausto Guillermo': ('F(60%), R(26%), C(7%)', '3yr'),
    'Brandi Cook': ('F(74%), R(4%), O(16%)', '5yr'),
    'Madeline Korbitz': ('F(42%), R(35%), I(13%)', '10yr'),
    'Patrick Morrissey': ('F(72%), R(16%), I(8%)', '5yr'),
    'David Schmidt': ('F(48%), I(33%), R(12%)', '5yr'),
    'Seamus Glavin': ('F(60%), R(25%), C(6%)', '2yr'),
    'Jessica Curry': ('F(59%), I(20%), R(16%)', '4yr'),
    'Bernie DeLacy': ('F(45%), I(30%), R(15%)', '25yr'),
    'Bill Odonnell': ('F(66%), R(18%), C(10%)', '7yr'),
    'Tom Kane': ('F(44%), I(24%), R(22%)', '18yr'),
    'Dave Lubin': ('F(57%), R(29%), I(10%)', '4yr'),
    'Andrea Biberger': ('F(72%), R(18%), C(8%)', '6yr'),
    'Tim Kirkey': ('F(54%), R(23%), I(17%)', '2yr'),
    'Vincent Scannapieco': ('F(43%), R(30%), C(16%)', '4yr'),
    'Phil Cerami': ('F(57%), R(23%), C(7%)', '4yr'),
    'Joshua Glavin': ('F(38%), C(30%), R(23%)', '7yr'),
    'Mary Ruth Bascelli': ('F(44%), R(22%), C(18%)', '5yr'),
    'Eric Schwarzel': ('F(35%), I(33%), R(19%)', '3yr'),
    'Courtney Blue': ('F(47%), I(37%), R(12%)', '3yr'),
    'Gabrielle Santorio': ('F(71%), R(10%), O(11%)', '6yr'),
    'Timothy Reilly': ('F(39%), I(26%), R(26%)', '3yr'),
    'Meghan Weaver': ('F(67%), R(22%), O(5%)', '4yr'),
    'Dan Rea': ('F(100%)', '7yr'),
    'Drew Brady': ('F(69%), R(18%), C(8%)', '4yr'),
    'Kara Klund': ('C(36%), F(31%), R(22%)', '2yr'),
    'Jim Rygiel': ('I(63%), R(19%), F(15%)', '2yr'),
    'Michael Woessner': ('F(78%), R(11%), C(6%)', '4yr'),
    'Danielle Penge': ('F(65%), R(17%), C(11%)', '3yr'),
    'Jaclyn Pacitti': ('F(64%), R(19%), C(11%)', '3yr'),
    'Michael Flon': ('F(61%), R(16%), I(12%)', '2yr'),
    'Michele Kenyon': ('F(66%), R(20%), I(12%)', '3yr'),
    'Mark DeBiase': ('Multi-type', '2yr'),
    'Chris Newell': ('Multi-type', '2yr'),
    'Fred Mischler': ('I(45%), R(36%), F(18%)', '30yr'),
    'Denis Kennedy': ('F(90%), R(6%)', '28yr'),
}

for row in range(5, ws_dir.max_row + 1):
    name = ws_dir.cell(row, 1).value
    if name and name in rep_extra:
        inst, yrs = rep_extra[name]
        sc(ws_dir, row, 8, inst)
        sc(ws_dir, row, 9, yrs)

ws_dir.column_dimensions['H'].width = 28
ws_dir.column_dimensions['I'].width = 12

wb.save(r'C:\tmp\tsi-sales-rep-benchmarks.xlsx')
print('Updated with Yr 1-5 Progression dashboard and enriched Rep Directory')
