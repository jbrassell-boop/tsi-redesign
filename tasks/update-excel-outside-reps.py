import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(r'C:\tmp\tsi-sales-rep-benchmarks.xlsx')

hdr_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='143B87')
normal_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', bold=True, size=10)
section_font = Font(name='Arial', bold=True, size=12, color='143B87')
title_font = Font(name='Arial', bold=True, size=16, color='143B87')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
money_fmt = '$#,##0;($#,##0);"-"'
pct_fmt = '0.0%'
green_fill = PatternFill('solid', fgColor='C6EFCE')
light_fill = PatternFill('solid', fgColor='D6E4F0')
yellow_fill = PatternFill('solid', fgColor='FFFF00')
red_fill = PatternFill('solid', fgColor='FFC7CE')
input_fill = PatternFill('solid', fgColor='FFFF00')
thick_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def sc(ws, r, c, val=None, fmt=None, bold=False):
    cell = ws.cell(row=r, column=c)
    if val is not None: cell.value = val
    cell.font = bold_font if bold else normal_font
    cell.border = thin_border
    if fmt: cell.number_format = fmt
    return cell

# Remove old Yr 1-5 Progression sheet and replace
if 'Outside Rep Benchmark' in wb.sheetnames:
    del wb['Outside Rep Benchmark']

ws = wb.create_sheet('Outside Rep Benchmark', 0)
ws.sheet_properties.tabColor = '143B87'

ws['A1'] = 'Outside Territory Rep Benchmark (5-Year Lookback)'
ws['A1'].font = Font(name='Arial', bold=True, size=16, color='143B87')
ws.merge_cells('A1:K1')
ws['A2'] = '8 confirmed outside territory reps from TSI history | Invoice-based data | FFS only (no contract revenue)'
ws['A2'].font = Font(name='Arial', italic=True, size=11, color='666666')
ws.merge_cells('A2:K2')

# ============================================================
# Section 1: Individual Rep Year-by-Year
# ============================================================
r = 4
ws.cell(r, 1, 'INDIVIDUAL REP PROGRESSION').font = section_font

# Per-rep data: (name, territory, started, [(yr, total, netnew, inherited, nn_clients, inh_clients, flex%, rigid%, camera%)])
rep_data = [
    ('Lindsey Davis', 'MD', 2005, [
        (1, 655197, 67046, 588151, 5, 14, 56, 11, 13),
        (2, 781567, 66140, 715426, 5, 13, 53, 22, 12),
        (3, 1623908, 954785, 669123, 5, 11, 54, 29, 11),
        (4, 1629390, 1048828, 580561, 6, 8, 53, 36, 6),
        (5, 1455687, 1083019, 372668, 5, 9, 54, 34, 8),
    ]),
    ('Eric Schwarzel', 'PA/NJ', 2007, [
        (1, 327210, 10323, 316887, 4, 29, 36, 16, 14),
        (2, 784957, 25031, 759926, 4, 34, 32, 20, 13),
        (3, 151231, 8225, 143006, 4, 25, 42, 25, 9),
    ]),
    ('Courtney Blue', 'OR/TX', 2007, [
        (1, 444155, 48116, 396039, 10, 10, 39, 17, 5),
        (2, 478997, 58620, 420377, 5, 6, 48, 9, 3),
        (3, 327543, 55130, 272413, 5, 6, 61, 8, 3),
    ]),
    ('Danielle Penge', 'OR/NJ', 2007, [
        (1, 386599, 21221, 365378, 4, 23, 64, 14, 15),
        (2, 348274, 9452, 338823, 4, 16, 63, 21, 8),
        (3, 142782, 0, 142782, 0, 5, 100, 0, 0),
    ]),
    ('Michael Woessner', 'NJ/OR', 2005, [
        (1, 223554, 852, 222702, 2, 14, 75, 10, 6),
        (2, 452289, 4957, 447332, 3, 20, 78, 11, 6),
        (3, 209953, 5708, 204244, 2, 12, 80, 13, 6),
    ]),
    ('Kara Klund', 'IL/TX', 2006, [
        (1, 562615, 2395, 560220, 2, 18, 33, 25, 34),
        (2, 343497, 690, 342807, 2, 12, 29, 18, 38),
    ]),
    ('Bernie DeLacy', 'FL/PA', 2001, [
        (1, 133326, 43471, 89855, 18, 24, 91, 2, 5),
        (2, 11301, 4965, 6336, 2, 5, 46, 46, 0),
    ]),
    ('Tom Kane', 'TX/PA', 2015, [
        (1, 306662, 0, 306662, 0, 18, 44, 22, 7),
        (2, 372892, 0, 372892, 0, 21, 44, 22, 7),
        (3, 589936, 0, 589936, 0, 22, 44, 22, 7),
        (4, 484868, 0, 484868, 0, 24, 44, 22, 7),
        (5, 293183, 0, 293183, 0, 22, 44, 22, 7),
    ]),
]

headers = ['Rep Name', 'Territory', 'Started', 'Year', 'Total Revenue', 'Net New Rev',
           'Inherited Rev', 'Net New %', 'NN Clients', 'Inh Clients', 'Total Clients',
           'Flex %', 'Rigid %', 'Camera %']
r += 1
for c, h in enumerate(headers, 1): ws.cell(r, c, h)
style_header(ws, r, 14)

current_row = r + 1
for rep_name, territory, started, years in rep_data:
    first_row = current_row
    for yr, total, nn, inh, nn_c, inh_c, flex, rigid, cam in years:
        if current_row == first_row:
            sc(ws, current_row, 1, rep_name, bold=True)
            sc(ws, current_row, 2, territory)
            sc(ws, current_row, 3, started)
            ws.cell(current_row, 3).number_format = '0'
        sc(ws, current_row, 4, f'Year {yr}')
        sc(ws, current_row, 5, total, money_fmt)
        sc(ws, current_row, 6, nn, money_fmt)
        sc(ws, current_row, 7, inh, money_fmt)
        sc(ws, current_row, 8, nn / total if total > 0 else 0, pct_fmt)
        sc(ws, current_row, 9, nn_c)
        sc(ws, current_row, 10, inh_c)
        sc(ws, current_row, 11, nn_c + inh_c)
        sc(ws, current_row, 12, flex / 100, pct_fmt)
        sc(ws, current_row, 13, rigid / 100, pct_fmt)
        sc(ws, current_row, 14, cam / 100, pct_fmt)
        if (current_row - r - 1) % 2 == 0:
            for c in range(1, 15): ws.cell(current_row, c).fill = light_fill
        current_row += 1
    current_row += 1  # gap between reps

# ============================================================
# Section 2: Averages — Outside Reps Only
# ============================================================
current_row += 1
ws.cell(current_row, 1, 'OUTSIDE TERRITORY REP AVERAGES').font = section_font
ws.merge_cells(f'A{current_row}:K{current_row}')
current_row += 1

avg_headers = ['', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5']
for c, h in enumerate(avg_headers, 1): ws.cell(current_row, c, h)
style_header(ws, current_row, 6)
current_row += 1

# Averages from the SQL output
avg_data = [
    ('Total Revenue (avg)', [303667, 457269, 491083, 814695, 1455687], money_fmt),
    ('  Net New Revenue', [21530, 24265, 204770, 524414, 1083019], money_fmt),
    ('  Inherited Revenue', [282137, 433004, 286314, 290281, 372668], money_fmt),
    ('Net New % of Total', [0.07, 0.05, 0.42, 0.64, 0.74], pct_fmt),
    ('Avg Unique Clients', [20, 19, 15, 8, 14], '#,##0'),
    ('FFS Revenue', [303667, 457269, 491083, 814695, 1455687], money_fmt),
    ('Contract Revenue', [0, 0, 0, 0, 0], money_fmt),
    ('# Reps in Sample', [9, 7, 5, 2, 1], '#,##0'),
]

for label, vals, fmt in avg_data:
    sc(ws, current_row, 1, label, bold='Total' in label or 'Net New %' in label or 'Avg' in label)
    for c, v in enumerate(vals, 2): sc(ws, current_row, c, v, fmt)
    if 'Total Revenue' in label:
        for c in range(1, 7): ws.cell(current_row, c).fill = light_fill
    if 'Net New %' in label:
        for c in range(1, 7): ws.cell(current_row, c).fill = green_fill
    current_row += 1

current_row += 1
ws.cell(current_row, 1, 'KEY INSIGHT: Outside territory reps inherit most of their Year 1-2 revenue.').font = Font(name='Arial', bold=True, size=11, color='FF0000')
ws.merge_cells(f'A{current_row}:K{current_row}')
current_row += 1
ws.cell(current_row, 1, 'Net new business doesn\'t meaningfully kick in until Year 3 (42%) and dominates by Year 4-5 (64-74%).').font = Font(name='Arial', italic=True, size=10, color='333333')
ws.merge_cells(f'A{current_row}:K{current_row}')
current_row += 1
ws.cell(current_row, 1, 'All 8 outside reps were 100% FFS — zero contract revenue. Contracts come from inside relationships.').font = Font(name='Arial', italic=True, size=10, color='333333')
ws.merge_cells(f'A{current_row}:K{current_row}')

# ============================================================
# Section 3: Comparison — Outside vs All Reps
# ============================================================
current_row += 3
ws.cell(current_row, 1, 'COMPARISON: OUTSIDE TERRITORY vs ALL REPS').font = section_font
ws.merge_cells(f'A{current_row}:F{current_row}')
current_row += 1
for c, h in enumerate(avg_headers, 1): ws.cell(current_row, c, h)
style_header(ws, current_row, 6)
current_row += 1

comp_data = [
    ('Outside Reps — Total Rev', [303667, 457269, 491083, 814695, 1455687], green_fill),
    ('All Reps (109) — Total Rev', [231914, 274479, 340148, 377677, 492651], light_fill),
    ('Outside Reps — Net New', [21530, 24265, 204770, 524414, 1083019], green_fill),
    ('All Reps (109) — Net New', [44085, 74028, 146008, 176736, 257577], light_fill),
]
for label, vals, fill in comp_data:
    sc(ws, current_row, 1, label, bold=True)
    for c, v in enumerate(vals, 2): sc(ws, current_row, c, v, money_fmt)
    for c in range(1, 7): ws.cell(current_row, c).fill = fill
    current_row += 1

current_row += 1
ws.cell(current_row, 1, 'Outside territory reps generate MORE total revenue (they get big inherited books) but LESS net new early on.').font = Font(name='Arial', italic=True, size=10)
ws.merge_cells(f'A{current_row}:K{current_row}')
current_row += 1
ws.cell(current_row, 1, 'However, by Year 3-5 their net new revenue dwarfs the all-rep average — they find big accounts.').font = Font(name='Arial', italic=True, size=10)
ws.merge_cells(f'A{current_row}:K{current_row}')

# Update Revenue Planner with outside rep numbers
ws_plan = wb['Revenue Planner']
# Update the assumption cells (row 12) with outside rep averages
outside_yr_vals = [303667, 457269, 491083, 814695, 1455687]
for c, v in enumerate(outside_yr_vals, 2):
    cell = ws_plan.cell(12, c)
    cell.value = v

# Column widths
for c in range(1, 15):
    ws.column_dimensions[get_column_letter(c)].width = 14
ws.column_dimensions['A'].width = 28
ws.column_dimensions['E'].width = 16
ws.column_dimensions['F'].width = 16
ws.column_dimensions['G'].width = 16

wb.save(r'C:\tmp\tsi-sales-rep-benchmarks-v2.xlsx')
print('Updated with Outside Rep Benchmark sheet')
print('Sheets:', [s.title for s in wb.worksheets])
