import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.load_workbook(r'C:\tmp\tsi-sales-rep-benchmarks.xlsx')

hdr_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='143B87')
normal_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', bold=True, size=10)
section_font = Font(name='Arial', bold=True, size=12, color='143B87')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
money_fmt = '$#,##0;($#,##0);"-"'
outside_fill = PatternFill('solid', fgColor='C6EFCE')

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def sc(ws, row, col, val=None, fmt=None):
    cell = ws.cell(row=row, column=col)
    if val is not None:
        cell.value = val
    cell.font = normal_font
    cell.border = thin_border
    if fmt:
        cell.number_format = fmt
    return cell

# Rep geographic data: (name, total_rev, primary_state, top_states, type_guess)
# type: O=outside territory, I=inside/phone, P=program/channel
geo_data = [
    ('Brian Kenney', 21312950, 'NY', 'NY, KY, IL', 'O', '2011'),
    ('Debbie Hightower', 19072425, 'GA', 'GA, TN, TX', 'O', '1999'),
    ('John Luedke', 12596333, 'OH', 'OH, TX, NC', 'O', '2005'),
    ('Gina Ellis', 7701908, 'TX', 'TX, NY, CA', 'O', '2008'),
    ('Lindsey Davis', 6548594, 'MD', 'MD, NY, CA', 'O', '2005'),
    ('Ryan George', 4607281, 'SD', 'SD, CA, TX', 'O', '2017'),
    ('Maria Mercanti', 4439325, 'VA', 'VA, TX, PA', 'O', '2017'),
    ('Jim Hueston', 3334808, 'TX', 'TX, WI, PA', 'O', '2006'),
    ('Ryan Gades', 3175278, 'CA', 'CA, TX', 'O', '2007'),
    ('Fausto Guillermo', 2947809, 'CA', 'CA, OR, IN', 'O', '2021'),
    ('Brandi Cook', 2868185, 'PA', 'PA, OH, TX', 'O', '2021'),
    ('Madeline Korbitz', 2600469, 'CT', 'CT, TX, LA', 'O', '2021'),
    ('Patrick Morrissey', 2506558, 'PA', 'PA, MI', 'O', '1998'),
    ('David Schmidt', 2376615, 'IL', 'IL, WI, IN', 'O', '2000'),
    ('Seamus Glavin', 1928017, 'CA', 'CA, AZ, VA', 'O', '2024'),
    ('Jessica Curry', 1924434, 'MD', 'MD, PA, WV', 'O', '2001'),
    ('Bernie DeLacy', 1907667, 'FL', 'FL, PA, NJ', 'O', '2001'),
    ('Bill Odonnell', 1750012, 'CA', 'CA', 'O', '2010'),
    ('Tom Kane', 1734537, 'TX', 'TX, PA, NJ', 'O', '2002'),
    ('Dave Lubin', 1581031, 'PA', 'PA, TX, DE', 'O', '1997'),
    ('Andrea Biberger', 1517813, 'TX', 'TX, PA, NY', 'O', '2012'),
    ('Tim Kirkey', 1457995, 'NY', 'NY, MD, CA', 'O', '2011'),
    ('Vincent Scannapieco', 1447615, 'LA', 'LA, OK', 'O', '2011'),
    ('Phil Cerami', 1310184, 'CA', 'CA', 'O', '2013'),
    ('Joshua Glavin', 1334396, 'OK', 'OK, LA, IL', 'O', '2004'),
    ('Mary Ruth Bascelli', 1283218, 'AR', 'AR, TX, MO', 'O', '2007'),
    ('Eric Schwarzel', 1263399, 'PA', 'PA, NJ, DE', 'O', '2007'),
    ('Courtney Blue', 1250695, 'OR', 'OR, TX, WI', 'O', '2007'),
    ('Gabrielle Santorio', 1243031, 'OH', 'OH, AZ, TX', 'O', '2018'),
    ('Timothy Reilly', 1169645, 'TX', 'TX, NY, MI', 'O', '2023'),
    ('Meghan Weaver', 1147470, 'CA', 'CA, IN, NY', 'O', '2019'),
    ('Dan Rea', 1123017, 'CA', 'CA, PA', 'O', '1995'),
    ('Drew Brady', 987919, 'IL', 'IL, WI, TX', 'O', '2014'),
    ('Kara Klund', 906112, 'IL', 'IL, TX, OR', 'O', '2006'),
    ('Jim Rygiel', 886361, 'FL', 'FL, MD, WA', 'O', '2024'),
    ('Michael Woessner', 885795, 'NJ', 'NJ, OR, OH', 'O', '2005'),
    ('Danielle Penge', 877655, 'OR', 'OR, NJ, NY', 'O', '2007'),
    ('Jaclyn Pacitti', 832763, 'MI', 'MI, TX, IN', 'O', '2010'),
    ('Michael Flon', 771414, 'NJ', 'NJ, NY', 'O', '2002'),
    ('Michele Kenyon', 763599, 'NJ', 'NJ, NY, WV', 'O', '1999'),
    ('Mark DeBiase', 1603231, 'Multiple', 'Multi-state', 'O', '2024'),
    ('Chris Newell', 641908, 'Multiple', 'Multi-state', 'O', '2024'),
    ('Fred Mischler', 557283, 'PA', 'PA, NJ', 'O', '2024'),
    ('Denis Kennedy', 224915, 'TX', 'TX, OK, FL', 'O/I', '1997'),
]

ws7 = wb.create_sheet('Rep Directory')
ws7.sheet_properties.tabColor = '548235'

ws7['A1'] = 'Sales Rep Directory — Territory & Revenue'
ws7['A1'].font = Font(name='Arial', bold=True, size=14, color='143B87')
ws7.merge_cells('A1:G1')
ws7['A2'] = 'Geographic data from invoice bill-to addresses. Highlights = outside territory reps.'
ws7['A2'].font = Font(name='Arial', italic=True, size=10, color='666666')
ws7.merge_cells('A2:G2')

headers = ['Rep Name', 'Primary Territory', 'Top States', 'Career Revenue', 'Started', 'Type', 'Notes']
for c, h in enumerate(headers, 1):
    ws7.cell(4, c, h)
style_header_row(ws7, 4, 7)

for i, (name, rev, primary, states, rtype, started) in enumerate(geo_data):
    row = 5 + i
    sc(ws7, row, 1, name).font = bold_font
    sc(ws7, row, 2, primary)
    sc(ws7, row, 3, states)
    sc(ws7, row, 4, rev, money_fmt)
    sc(ws7, row, 5, int(started))
    ws7.cell(row, 5).number_format = '0'
    sc(ws7, row, 6, 'Outside' if 'O' in rtype else 'Inside')
    sc(ws7, row, 7, '')
    if 'O' in rtype:
        for c in range(1, 8):
            ws7.cell(row, c).fill = outside_fill

widths = [22, 16, 20, 16, 10, 10, 30]
for c, w in enumerate(widths, 1):
    ws7.column_dimensions[chr(64+c)].width = w

# Add legend
lr = 5 + len(geo_data) + 2
ws7.cell(lr, 1, 'Legend:').font = bold_font
ws7.cell(lr+1, 1, 'Green = Outside Territory Rep (boots on the ground)')
ws7.cell(lr+1, 1).fill = outside_fill
ws7.cell(lr+2, 1, 'These are the reps most relevant for benchmarking new territory hires.')
ws7.cell(lr+2, 1).font = Font(name='Arial', italic=True, size=10)

wb.save(r'C:\tmp\tsi-sales-rep-benchmarks.xlsx')
print('Updated with Rep Directory sheet')
