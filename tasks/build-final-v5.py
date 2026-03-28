import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
hdr_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='143B87')
normal_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', bold=True, size=10)
section_font = Font(name='Arial', bold=True, size=12, color='143B87')
title_font = Font(name='Arial', bold=True, size=16, color='143B87')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
thick_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
money_fmt = '$#,##0;($#,##0);"-"'
pct_fmt = '0.0%'
green_fill = PatternFill('solid', fgColor='C6EFCE')
light_fill = PatternFill('solid', fgColor='D6E4F0')
yellow_fill = PatternFill('solid', fgColor='FFFF00')
red_fill = PatternFill('solid', fgColor='FFC7CE')
orange_fill = PatternFill('solid', fgColor='FCE4D6')
total_fill = PatternFill('solid', fgColor='143B87')
total_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')

def hdr(ws, row, max_col):
    for col in range(1, max_col + 1):
        c = ws.cell(row=row, column=col)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', wrap_text=True); c.border = thin_border

def sc(ws, r, c, val=None, fmt=None, bold=False):
    cell = ws.cell(row=r, column=c)
    if val is not None: cell.value = val
    cell.font = bold_font if bold else normal_font; cell.border = thin_border
    if fmt: cell.number_format = fmt
    return cell

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w

def total_row(ws, row, max_col, label='TOTAL'):
    ws.cell(row, 1, label)
    for c in range(1, max_col + 1):
        ws.cell(row, c).font = total_font; ws.cell(row, c).fill = total_fill; ws.cell(row, c).border = thin_border

# ============================================================
# SHEET 1: Revenue Planner (the tool Denis wants)
# ============================================================
ws = wb.active
ws.title = 'Revenue Planner'
ws.sheet_properties.tabColor = 'FF0000'

ws['A1'] = 'TSI Sales Revenue Planner'
ws['A1'].font = Font(name='Arial', bold=True, size=18, color='143B87')
ws.merge_cells('A1:I1')
ws['A2'] = 'Set your revenue goal. See current roster + projected new hires. Based on 25 years of invoice data.'
ws['A2'].font = Font(name='Arial', italic=True, size=11, color='666666')
ws.merge_cells('A2:I2')

# Revenue Goal
ws.cell(4, 1, 'REVENUE GOAL').font = section_font
sc(ws, 5, 1, 'Annual Revenue Target:', bold=True)
c = sc(ws, 5, 2, 20000000, money_fmt); c.fill = yellow_fill; c.border = thick_border; c.font = Font(name='Arial', bold=True, size=14)
sc(ws, 5, 3, '(change this)', bold=False).font = Font(name='Arial', italic=True, size=9, color='999999')

# Current Roster
ws.cell(7, 1, 'CURRENT REP ROSTER (2025 Revenue)').font = section_font
ws.merge_cells('A7:I7')
r = 8
headers = ['Rep Name', 'Type', '2024 FFS', '2024 Contract', '2024 Total', '2025 FFS', '2025 Contract', '2025 Total', 'Clients']
for c, h in enumerate(headers, 1): ws.cell(r, c, h)
hdr(ws, r, 9)

# Current reps data (from query results, sorted by 2025 total)
current_reps = [
    ('Brian Kenney', 'Inside', 2420708, 1138756, 4158828, 2550819, 76),
    ('Brandi Cook', 'Inside', 965564, 709126, 1152405, 804556, 22),
    ('Bernie DeLacy', 'Outside', 1109861, 592666, 1235032, 562849, 46),
    ('Seamus Glavin', 'Inside', 803922, 324594, 937107, 213479, 61),
    ('Debbie Hightower', 'Inside', 1353134, 116022, 2023520, 78162, 130),
    ('Jim Rygiel', 'Outside', 114669, 0, 646265, 124284, 11),
    ('Timothy Reilly', 'Inside', 625012, 73986, 396220, 86580, 47),
    ('Donald Haynes', 'Inside', 44317, 0, 207680, 0, 47),
    ('Ryan George', 'Inside', 190976, 168748, 116813, 0, 13),
    ('Mark DeBiase', 'Inside', 0, 0, 0, 0, 92),
]

# Also add BSC program revenue as a line
bsc_reps = [
    ('BSC Sole Flex', 'Program', 0, 567456, 0, 807336, 0),
    ('BSC CAP', 'Program', 0, 781600, 0, 562676, 0),
    ('BSC CAP - S', 'Program', 0, 384810, 0, 460852, 0),
    ('BSC CAP - M', 'Program', 0, 279282, 0, 331028, 0),
    ('EC Legacy', 'Program', 0, 66674, 0, 155587, 0),
]

row = r + 1
for name, rtype, ffs24, con24, ffs25, con25, clients in current_reps:
    sc(ws, row, 1, name, bold=True)
    sc(ws, row, 2, rtype)
    sc(ws, row, 3, ffs24, money_fmt)
    sc(ws, row, 4, con24, money_fmt)
    sc(ws, row, 5, f'=C{row}+D{row}', money_fmt)
    sc(ws, row, 6, ffs25, money_fmt)
    sc(ws, row, 7, con25, money_fmt)
    sc(ws, row, 8, f'=F{row}+G{row}', money_fmt)
    sc(ws, row, 9, clients)
    fill = green_fill if rtype == 'Outside' else light_fill if rtype == 'Inside' else orange_fill
    for c in range(1, 10): ws.cell(row, c).fill = fill
    row += 1

# BSC programs
for name, rtype, ffs24, con24, ffs25, con25, clients in bsc_reps:
    sc(ws, row, 1, name, bold=True)
    sc(ws, row, 2, rtype)
    sc(ws, row, 3, ffs24, money_fmt)
    sc(ws, row, 4, con24, money_fmt)
    sc(ws, row, 5, f'=C{row}+D{row}', money_fmt)
    sc(ws, row, 6, ffs25, money_fmt)
    sc(ws, row, 7, con25, money_fmt)
    sc(ws, row, 8, f'=F{row}+G{row}', money_fmt)
    sc(ws, row, 9, clients)
    for c in range(1, 10): ws.cell(row, c).fill = orange_fill
    row += 1

# Subtotal current
current_first = r + 1
current_last = row - 1
total_row(ws, row, 9, 'CURRENT ROSTER TOTAL')
for col in range(3, 9):
    cl = get_column_letter(col)
    ws.cell(row, col, f'=SUM({cl}{current_first}:{cl}{current_last})')
    ws.cell(row, col).number_format = money_fmt
    ws.cell(row, col).font = total_font; ws.cell(row, col).fill = total_fill
current_total_row = row
row += 2

# New Hires section
ws.cell(row, 1, 'NEW HIRES (projected using historical averages)').font = section_font
ws.merge_cells(f'A{row}:I{row}')
row += 1
ws.cell(row, 1, 'Edit yellow cells. Revenue auto-fills from historical benchmarks.').font = Font(name='Arial', italic=True, size=9, color='666666')
row += 1

# Assumptions row (hidden reference)
assume_row = row
sc(ws, row, 1, 'Yr1 Outside Avg:', bold=True)
sc(ws, row, 2, 348374, money_fmt).fill = yellow_fill
sc(ws, row, 3, 'Yr1 Inside Avg:', bold=True)
sc(ws, row, 4, 202000, money_fmt).fill = yellow_fill
sc(ws, row, 5, 'YoY Growth:', bold=True)
sc(ws, row, 6, 0.30, pct_fmt).fill = yellow_fill
sc(ws, row, 7, '(30% avg yr-over-yr)', bold=False).font = Font(name='Arial', italic=True, size=9, color='999999')
row += 1

# New hire headers
for c, h in enumerate(['New Hire Name', 'Type', 'Start Year', 'Year 1 Rev', 'Year 2 Rev', 'Year 3 Rev', 'Year 4 Rev', 'Year 5 Rev', '5-Yr Total'], 1):
    ws.cell(row, c, h)
hdr(ws, row, 9)
new_hire_hdr = row
row += 1

# 8 new hire rows
for i in range(8):
    sc(ws, row, 1, f'New Hire {i+1}').fill = yellow_fill
    # Type dropdown placeholder
    c = sc(ws, row, 2, 'Outside' if i < 3 else 'Inside')
    c.fill = yellow_fill
    sc(ws, row, 3, 2026 + (i // 3)).fill = yellow_fill
    ws.cell(row, 3).number_format = '0'
    # Year 1-5 revenue formulas
    # If type=Outside use B assume_row, else D assume_row, then compound growth
    for yr in range(1, 6):
        col = 3 + yr
        if yr == 1:
            sc(ws, row, col, f'=IF(B{row}="Outside",$B${assume_row},$D${assume_row})', money_fmt)
        else:
            prev_col = get_column_letter(col - 1)
            sc(ws, row, col, f'={prev_col}{row}*(1+$F${assume_row})', money_fmt)
    sc(ws, row, 9, f'=SUM(D{row}:H{row})', money_fmt)
    if i % 2 == 0:
        for c in range(1, 10): ws.cell(row, c).fill = PatternFill('solid', fgColor='E2EFDA') if 'Outside' in str(ws.cell(row, 2).value) else light_fill
    row += 1

# New hire totals
new_first = new_hire_hdr + 1
new_last = row - 1
total_row(ws, row, 9, 'NEW HIRES TOTAL')
for col in range(4, 10):
    cl = get_column_letter(col)
    ws.cell(row, col, f'=SUM({cl}{new_first}:{cl}{new_last})')
    ws.cell(row, col).number_format = money_fmt
    ws.cell(row, col).font = total_font; ws.cell(row, col).fill = total_fill
new_total_row = row
row += 2

# Summary
ws.cell(row, 1, 'SUMMARY').font = Font(name='Arial', bold=True, size=14, color='143B87')
row += 1
for c, h in enumerate(['', '2025 Actual', 'Projected Year 1', 'Projected Year 2', 'Projected Year 3'], 1):
    ws.cell(row, c, h)
hdr(ws, row, 5)
row += 1

sc(ws, row, 1, 'Current Roster', bold=True)
sc(ws, row, 2, f'=H{current_total_row}', money_fmt)
sc(ws, row, 3, f'=H{current_total_row}', money_fmt)  # Assume flat
sc(ws, row, 4, f'=H{current_total_row}', money_fmt)
sc(ws, row, 5, f'=H{current_total_row}', money_fmt)
for c in range(1, 6): ws.cell(row, c).fill = light_fill
row += 1

sc(ws, row, 1, 'New Hires', bold=True)
sc(ws, row, 2, 0, money_fmt)
sc(ws, row, 3, f'=D{new_total_row}', money_fmt)
sc(ws, row, 4, f'=E{new_total_row}', money_fmt)
sc(ws, row, 5, f'=F{new_total_row}', money_fmt)
for c in range(1, 6): ws.cell(row, c).fill = green_fill
row += 1

total_row(ws, row, 5, 'PROJECTED TOTAL')
for col in range(2, 6):
    cl = get_column_letter(col)
    ws.cell(row, col, f'={cl}{row-2}+{cl}{row-1}')
    ws.cell(row, col).number_format = money_fmt
    ws.cell(row, col).font = total_font; ws.cell(row, col).fill = total_fill
proj_total_row = row
row += 1

sc(ws, row, 1, 'Revenue Goal', bold=True)
for col in range(2, 6):
    sc(ws, row, col, f'=$B$5', money_fmt)
ws.cell(row, 1).fill = yellow_fill
row += 1

sc(ws, row, 1, 'GAP TO GOAL', bold=True).font = Font(name='Arial', bold=True, size=12, color='FF0000')
for col in range(2, 6):
    cl = get_column_letter(col)
    cell = sc(ws, row, col, f'={cl}{proj_total_row}-{cl}{row-1}', money_fmt)
    cell.font = Font(name='Arial', bold=True, size=12, color='FF0000')

set_widths(ws, [22, 10, 16, 16, 16, 16, 16, 16, 12])

# ============================================================
# SHEET 2: Inside vs Outside History
# ============================================================
ws2 = wb.create_sheet('Inside vs Outside')
ws2.sheet_properties.tabColor = '143B87'

ws2['A1'] = 'Historical Comparison: Inside vs Outside Territory Reps'
ws2['A1'].font = title_font
ws2.merge_cells('A1:G1')
ws2['A2'] = '8 outside reps + 101 inside reps | FFS + Contract Revenue | By tenure year'
ws2['A2'].font = Font(name='Arial', italic=True, size=10, color='666666')

r = 4
ws2.cell(r, 1, 'TOTAL REVENUE (FFS + Contract) BY TENURE YEAR').font = section_font
r += 1
for c, h in enumerate(['', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'], 1):
    ws2.cell(r, c, h)
hdr(ws2, r, 6)

# Corrected data with contract revenue included
comp = [
    ('Outside Territory (8 reps)', [348374, 404261, 491083, 814695, 1455687], green_fill),
    ('Inside / Phone (101+ reps)', [136078, 218961, 301268, 312861, 390956], light_fill),
    ('Outside Premium %', None, orange_fill),
]
for i, (label, vals, fill) in enumerate(comp):
    row = r + 1 + i
    sc(ws2, row, 1, label, bold=True)
    if vals:
        for c, v in enumerate(vals, 2): sc(ws2, row, c, v, money_fmt)
    else:
        for c in range(2, 7):
            sc(ws2, row, c, f'={get_column_letter(c)}{r+1}/{get_column_letter(c)}{r+2}-1', pct_fmt)
    for c in range(1, 7): ws2.cell(row, c).fill = fill

# FFS vs Contract breakdown
r += 6
ws2.cell(r, 1, 'REVENUE BREAKDOWN: FFS vs CONTRACT').font = section_font
r += 1
for c, h in enumerate(['', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'], 1):
    ws2.cell(r, c, h)
hdr(ws2, r, 6)

breakdown = [
    ('Outside — FFS', [348374, 404261, 491083, 814695, 1455687], green_fill),
    ('Outside — Contract', [0, 0, 0, 0, 0], green_fill),
    ('Inside — FFS', [127025, 174756, 223942, 239933, 302433], light_fill),
    ('Inside — Contract', [9053, 44204, 77325, 72928, 88523], light_fill),
]
for i, (label, vals, fill) in enumerate(breakdown):
    row = r + 1 + i
    sc(ws2, row, 1, label, bold=True)
    for c, v in enumerate(vals, 2): sc(ws2, row, c, v, money_fmt)
    for c in range(1, 7): ws2.cell(row, c).fill = fill

# Net New
r += 7
ws2.cell(r, 1, 'NET NEW REVENUE (self-sourced)').font = section_font
r += 1
for c, h in enumerate(['', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'], 1):
    ws2.cell(r, c, h)
hdr(ws2, r, 6)

nn = [
    ('Outside — Net New', [24178, 21232, 204770, 524414, 1083019]),
    ('Outside — Net New %', [0.07, 0.05, 0.42, 0.64, 0.74]),
    ('Inside — Net New', [28212, 57618, 95843, 102939, 129510]),
    ('Inside — Net New %', [0.21, 0.26, 0.32, 0.33, 0.33]),
]
fills2 = [green_fill, green_fill, light_fill, light_fill]
for i, (label, vals) in enumerate(nn):
    row = r + 1 + i
    sc(ws2, row, 1, label, bold=True)
    fmt = pct_fmt if '%' in label else money_fmt
    for c, v in enumerate(vals, 2): sc(ws2, row, c, v, fmt)
    for c in range(1, 7): ws2.cell(row, c).fill = fills2[i]

# Key insights
r += 7
ws2.cell(r, 1, 'KEY INSIGHTS').font = Font(name='Arial', bold=True, size=14, color='FF0000')
insights = [
    'Outside reps generate 2.6x more revenue in Year 1 ($348K vs $136K) — mostly from inherited books.',
    'Inside reps add contract revenue earlier ($9K yr1, $44K yr2, $77K yr3). Outside reps = 100% FFS.',
    'Inside reps source more net new clients faster (22% in yr1 vs 7% for outside).',
    'Outside reps hit the big numbers by Year 4-5 ($800K-$1.5M) once their territory matures.',
    'Contract revenue is an inside rep play — driven by relationship depth, not territory coverage.',
    '53% of inside reps leave within 3 years. The ones who stay become $1M+ producers.',
    'The real ROI question: 1 outside rep at $800K/yr or 2 inside reps at $400K/yr each?',
]
for i, text in enumerate(insights):
    ws2.cell(r+1+i, 1, f'  {i+1}. {text}').font = Font(name='Arial', size=10)
    ws2.merge_cells(f'A{r+1+i}:F{r+1+i}')

set_widths(ws2, [34, 14, 14, 14, 14, 14])

# ============================================================
# SHEET 3: Outside Rep Detail
# ============================================================
ws3 = wb.create_sheet('Outside Reps')
ws3.sheet_properties.tabColor = '70AD47'
ws3['A1'] = 'Outside Territory Rep Profiles'
ws3['A1'].font = title_font

rep_data = [
    ('Lindsey Davis', 'MD/DC', '2005-2010', 6548594, 0, 19, '54% Flex, 28% Rigid'),
    ('Eric Schwarzel', 'PA/NJ/DE', '2007-2010', 1263399, 0, 38, '35% Flex, 19% Rigid'),
    ('Courtney Blue', 'OR/TX', '2007-2010', 1250695, 0, 20, '47% Flex, 37% Instrument'),
    ('Danielle Penge', 'OR/NJ', '2007-2010', 877655, 0, 27, '65% Flex, 17% Rigid'),
    ('Michael Woessner', 'NJ/OR', '2005-2009', 885795, 0, 23, '78% Flex, 11% Rigid'),
    ('Kara Klund', 'IL/TX', '2006-2008', 906112, 0, 20, '33% Flex, 25% Rigid, 34% Camera'),
    ('Melissa Fox', 'MD/DC/VA', '2016-2018', 407129, 319586, 9, '72% Flex, 19% Rigid'),
    ('Bernie DeLacy', 'FL/PA', '~2023+', 371597, 1870094, 46, 'Mixed (recent stint)'),
]

for c, h in enumerate(['Rep', 'Territory', 'Active', 'Career FFS', 'Career Contract', 'Career Total', 'Clients', 'Instrument Mix'], 1):
    ws3.cell(3, c, h)
hdr(ws3, 3, 8)

for i, (name, terr, active, ffs, con, clients, mix) in enumerate(rep_data):
    row = 4 + i
    sc(ws3, row, 1, name, bold=True)
    sc(ws3, row, 2, terr)
    sc(ws3, row, 3, active)
    sc(ws3, row, 4, ffs, money_fmt)
    sc(ws3, row, 5, con, money_fmt)
    sc(ws3, row, 6, f'=D{row}+E{row}', money_fmt)
    sc(ws3, row, 7, clients)
    sc(ws3, row, 8, mix)
    if i % 2 == 0:
        for c in range(1, 9): ws3.cell(row, c).fill = light_fill

set_widths(ws3, [20, 14, 14, 16, 16, 16, 10, 34])

# ============================================================
# SHEET 4: Retention
# ============================================================
ws4 = wb.create_sheet('Retention')
ws4.sheet_properties.tabColor = 'FFC000'
ws4['A1'] = 'Inside Rep Retention & Tenure'
ws4['A1'].font = title_font

for c, h in enumerate(['Tenure', '# Reps', '% of Total', 'Cumulative Gone', 'What It Means'], 1):
    ws4.cell(3, c, h)
hdr(ws4, 3, 5)

tenure = [
    ('1-2 years', 25, 0.21, 0.21, 'Gone before breakeven — net loss', red_fill),
    ('2-3 years', 38, 0.32, 0.53, '53% gone by year 3 — barely past ramp', orange_fill),
    ('3-4 years', 22, 0.18, 0.71, 'Starting to produce, then leave', PatternFill('solid', fgColor='FFEB9C')),
    ('4-5 years', 12, 0.10, 0.81, 'Solid but miss the peak years', light_fill),
    ('5+ years', 23, 0.19, 1.00, 'The money makers — $3.5M avg career', green_fill),
]
for i, (t, count, pct, cum, note, fill) in enumerate(tenure):
    row = 4 + i
    sc(ws4, row, 1, t, bold=True)
    sc(ws4, row, 2, count)
    sc(ws4, row, 3, pct, pct_fmt)
    sc(ws4, row, 4, cum, pct_fmt)
    sc(ws4, row, 5, note)
    for c in range(1, 6): ws4.cell(row, c).fill = fill

set_widths(ws4, [14, 10, 12, 16, 50])

# ============================================================
# SHEET 5: Company Totals
# ============================================================
ws5 = wb.create_sheet('Company Totals')
ws5.sheet_properties.tabColor = '4472C4'
ws5['A1'] = 'TSI Revenue Summary'
ws5['A1'].font = title_font

for c, h in enumerate(['Year', 'FFS Revenue', 'Contract Revenue', 'Total Revenue', 'Contract %'], 1):
    ws5.cell(3, c, h)
hdr(ws5, 3, 5)

totals = [
    (2024, 8958775, 5323596),
    (2025, 11771287, 6613925),
]
for i, (yr, ffs, con) in enumerate(totals):
    row = 4 + i
    sc(ws5, row, 1, yr); ws5.cell(row, 1).number_format = '0'
    sc(ws5, row, 2, ffs, money_fmt)
    sc(ws5, row, 3, con, money_fmt)
    sc(ws5, row, 4, f'=B{row}+C{row}', money_fmt)
    sc(ws5, row, 5, f'=C{row}/D{row}', pct_fmt)

ws5.cell(7, 1, 'Contract revenue = 37% of total (2025). This is NOT captured in FFS-only rep benchmarks.').font = Font(name='Arial', bold=True, size=10, color='FF0000')
ws5.merge_cells('A7:E7')

set_widths(ws5, [10, 16, 16, 16, 12])

# ============================================================
# Save
# ============================================================
out = r'C:\tmp\tsi-sales-rep-benchmarks-v5.xlsx'
wb.save(out)
print(f'Saved to {out}')
print('Sheets:', [s.title for s in wb.worksheets])
