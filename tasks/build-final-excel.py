import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
hdr_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='143B87')
normal_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', bold=True, size=10)
section_font = Font(name='Arial', bold=True, size=12, color='143B87')
title_font = Font(name='Arial', bold=True, size=16, color='143B87')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
thick_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
money_fmt = '$#,##0;($#,##0);"-"'
pct_fmt = '0.0%'
green_fill = PatternFill('solid', fgColor='C6EFCE')
light_fill = PatternFill('solid', fgColor='D6E4F0')
yellow_fill = PatternFill('solid', fgColor='FFFF00')
red_fill = PatternFill('solid', fgColor='FFC7CE')
orange_fill = PatternFill('solid', fgColor='FCE4D6')
total_fill = PatternFill('solid', fgColor='143B87')
total_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')

def hdr(ws, row, max_col, fill=None):
    for col in range(1, max_col + 1):
        c = ws.cell(row=row, column=col)
        c.font = hdr_font
        c.fill = fill or hdr_fill
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = thin_border

def sc(ws, r, c, val=None, fmt=None, bold=False):
    cell = ws.cell(row=r, column=c)
    if val is not None: cell.value = val
    cell.font = bold_font if bold else normal_font
    cell.border = thin_border
    if fmt: cell.number_format = fmt
    return cell

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# SHEET 1: Inside vs Outside Comparison
# ============================================================
ws = wb.active
ws.title = 'Inside vs Outside'
ws.sheet_properties.tabColor = '143B87'

ws['A1'] = 'Sales Rep Expectations: Inside vs Outside Territory'
ws['A1'].font = Font(name='Arial', bold=True, size=18, color='143B87')
ws.merge_cells('A1:G1')
ws['A2'] = 'TSI Historical Data | 200K+ Invoices | 1994-2026 | 9 Outside Reps, 99 Inside Reps'
ws['A2'].font = Font(name='Arial', italic=True, size=11, color='666666')
ws.merge_cells('A2:G2')

# Revenue comparison
r = 4
ws.cell(r, 1, 'TOTAL REVENUE BY TENURE YEAR').font = section_font
r += 1
for c, h in enumerate(['', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'], 1):
    ws.cell(r, c, h)
hdr(ws, r, 6)

rows_data = [
    ('Outside Territory Rep', [300743, 375408, 528493, 800751, 890622], green_fill),
    ('Inside / Phone Rep', [194837, 223726, 279838, 292239, 408323], light_fill),
    ('Outside Premium', None, orange_fill),
]
for i, (label, vals, fill) in enumerate(rows_data):
    row = r + 1 + i
    sc(ws, row, 1, label, bold=True)
    if vals:
        for c, v in enumerate(vals, 2): sc(ws, row, c, v, money_fmt)
    else:
        # Formula row for premium %
        for c in range(2, 7):
            sc(ws, row, c, f'={get_column_letter(c)}{r+1}/{get_column_letter(c)}{r+2}-1', pct_fmt)
    for c in range(1, 7): ws.cell(row, c).fill = fill

# Sample sizes
sc(ws, r+4, 1, '# Reps in Sample', bold=True)
outside_n = [10, 10, 7, 4, 3]
inside_n = [99, 75, 42, 31, 20]
for c, v in enumerate(outside_n, 2): sc(ws, r+4, c, f'{v} outside')
sc(ws, r+5, 1, '', bold=True)
for c, v in enumerate(inside_n, 2): sc(ws, r+5, c, f'{v} inside')

# Net New Revenue
r += 8
ws.cell(r, 1, 'NET NEW REVENUE (self-sourced business)').font = section_font
r += 1
for c, h in enumerate(['', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'], 1):
    ws.cell(r, c, h)
hdr(ws, r, 6)

nn_data = [
    ('Outside — Net New $', [23337, 29596, 208109, 392760, 458719], green_fill),
    ('Outside — Net New %', [0.08, 0.08, 0.39, 0.49, 0.52], green_fill),
    ('Inside — Net New $', [42619, 72056, 125228, 137460, 214527], light_fill),
    ('Inside — Net New %', [0.22, 0.32, 0.45, 0.47, 0.53], light_fill),
]
for i, (label, vals, fill) in enumerate(nn_data):
    row = r + 1 + i
    sc(ws, row, 1, label, bold=True)
    fmt = pct_fmt if '%' in label else money_fmt
    for c, v in enumerate(vals, 2): sc(ws, row, c, v, fmt)
    for c in range(1, 7): ws.cell(row, c).fill = fill

# Unique clients
r += 7
ws.cell(r, 1, 'UNIQUE CLIENTS PER YEAR').font = section_font
r += 1
for c, h in enumerate(['', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'], 1):
    ws.cell(r, c, h)
hdr(ws, r, 6)

sc(ws, r+1, 1, 'Outside — Total Clients', bold=True)
for c, v in enumerate([21, 17, 23, 28, 45], 2): sc(ws, r+1, c, v)
for c in range(1, 7): ws.cell(r+1, c).fill = green_fill

sc(ws, r+2, 1, '  of which Net New', bold=True)
for c, v in enumerate([5, 4, 5, 11, 17], 2): sc(ws, r+2, c, v)

sc(ws, r+3, 1, 'Inside — Total Clients', bold=True)
for c, v in enumerate([21, 22, 24, 25, 31], 2): sc(ws, r+3, c, v)
for c in range(1, 7): ws.cell(r+3, c).fill = light_fill

sc(ws, r+4, 1, '  of which Net New', bold=True)
for c, v in enumerate([9, 10, 12, 14, 18], 2): sc(ws, r+4, c, v)

# Key insights
r += 7
ws.cell(r, 1, 'KEY INSIGHTS FOR DENIS').font = Font(name='Arial', bold=True, size=14, color='FF0000')
ws.merge_cells(f'A{r}:F{r}')
insights = [
    'Outside territory reps generate 54% more revenue in Year 1 and 118% more by Year 5.',
    'BUT inside reps source more net new business early ($43K vs $23K in Year 1).',
    'Outside reps coast on inherited books for 2 years — net new doesn\'t kick in until Year 3.',
    'Inside reps plateau around $400K by Year 5. Outside reps blow past $800K.',
    'Outside reps were 100% FFS (except Melissa Fox who landed a Walter Reed contract).',
    'The sweet spot: inside reps for quick net new + outside reps for long-term territory build.',
]
for i, text in enumerate(insights):
    ws.cell(r+1+i, 1, f'  {i+1}. {text}').font = Font(name='Arial', size=10)
    ws.merge_cells(f'A{r+1+i}:F{r+1+i}')

set_widths(ws, [36, 14, 14, 14, 14, 14])

# ============================================================
# SHEET 2: Outside Rep Detail
# ============================================================
ws2 = wb.create_sheet('Outside Rep Detail')
ws2.sheet_properties.tabColor = '70AD47'

ws2['A1'] = 'Outside Territory Rep — Individual Performance'
ws2['A1'].font = title_font
ws2.merge_cells('A1:J1')
ws2['A2'] = '9 confirmed outside territory reps identified by Joe Brassell'
ws2['A2'].font = Font(name='Arial', italic=True, size=10, color='666666')

# Per-rep data
rep_data = [
    ('Lindsey Davis', 'MD/DC', '2005', [
        (1, 655197, 67046, 588151, 5, 14, 19),
        (2, 781567, 66140, 715426, 5, 13, 18),
        (3, 1623908, 954785, 669123, 5, 11, 16),
        (4, 1629390, 1048828, 580561, 6, 8, 14),
        (5, 1455687, 1083019, 372668, 5, 9, 14),
    ]),
    ('Tom Kane', 'PA', '2015', [
        (1, 306662, 0, 306662, 0, 18, 18),
        (2, 372892, 0, 372892, 0, 21, 21),
        (3, 589936, 0, 589936, 0, 22, 22),
        (4, 484868, 0, 484868, 0, 24, 24),
        (5, 293183, 0, 293183, 0, 22, 22),
    ]),
    ('Eric Schwarzel', 'PA/NJ/DE', '2007', [
        (1, 327210, 10323, 316887, 4, 29, 33),
        (2, 784957, 25031, 759926, 4, 34, 38),
        (3, 151231, 8225, 143006, 4, 25, 29),
    ]),
    ('Courtney Blue', 'OR/TX', '2007', [
        (1, 444155, 48116, 396039, 10, 10, 20),
        (2, 478997, 58620, 420377, 5, 6, 11),
        (3, 327543, 55130, 272413, 5, 6, 11),
    ]),
    ('Danielle Penge', 'OR/NJ', '2007', [
        (1, 386599, 21221, 365378, 4, 23, 27),
        (2, 348274, 9452, 338823, 4, 16, 20),
        (3, 142782, 0, 142782, 0, 5, 5),
    ]),
    ('Michael Woessner', 'NJ/OR', '2005', [
        (1, 223554, 852, 222702, 2, 14, 16),
        (2, 452289, 4957, 447332, 3, 20, 23),
        (3, 209953, 5708, 204244, 2, 12, 14),
    ]),
    ('Kara Klund', 'IL/TX', '2006', [
        (1, 562615, 2395, 560220, 2, 18, 20),
        (2, 343497, 690, 342807, 2, 12, 14),
    ]),
    ('Melissa Fox', 'MD/DC/VA', '2016', [
        (1, 98666, 0, 98666, 0, 3, 3),
        (2, 281443, 0, 281443, 0, 9, 9),
    ]),
    ('Bernie DeLacy', 'FL/PA', '~2023 (recent stint)', [
        (1, 133326, 43471, 89855, 18, 24, 42),
        (2, 11301, 4965, 6336, 2, 5, 7),
    ]),
]

headers = ['Rep', 'Territory', 'Started', 'Year', 'Total Rev', 'Net New', 'Inherited', 'Net New %', 'NN Clients', 'Inh Clients', 'Total Clients']
r = 4
for c, h in enumerate(headers, 1): ws2.cell(r, c, h)
hdr(ws2, r, 11)

row = r + 1
for name, territory, started, years in rep_data:
    first = True
    for yr, total, nn, inh, nnc, inhc, allc in years:
        if first:
            sc(ws2, row, 1, name, bold=True)
            sc(ws2, row, 2, territory)
            sc(ws2, row, 3, started)
            first = False
        sc(ws2, row, 4, f'Year {yr}')
        sc(ws2, row, 5, total, money_fmt)
        sc(ws2, row, 6, nn, money_fmt)
        sc(ws2, row, 7, inh, money_fmt)
        sc(ws2, row, 8, nn/total if total > 0 else 0, pct_fmt)
        sc(ws2, row, 9, nnc)
        sc(ws2, row, 10, inhc)
        sc(ws2, row, 11, allc)
        row += 1
    row += 1  # gap

set_widths(ws2, [20, 14, 22, 10, 14, 14, 14, 12, 12, 12, 12])

# ============================================================
# SHEET 3: Revenue Planner
# ============================================================
ws3 = wb.create_sheet('Revenue Planner')
ws3.sheet_properties.tabColor = 'FF0000'

ws3['A1'] = 'Sales Revenue Headcount Planner'
ws3['A1'].font = Font(name='Arial', bold=True, size=18, color='143B87')
ws3.merge_cells('A1:H1')
ws3['A2'] = 'Change yellow cells to model different scenarios'
ws3['A2'].font = Font(name='Arial', italic=True, size=11, color='666666')

# Inputs
ws3.cell(4, 1, 'INPUTS').font = section_font
ws3.cell(5, 1, 'Number of Outside Reps:').font = bold_font
c = sc(ws3, 5, 2, 3)
c.fill = yellow_fill; c.border = thick_border; c.font = Font(name='Arial', bold=True, size=12)

ws3.cell(6, 1, 'Number of Inside Reps:').font = bold_font
c = sc(ws3, 6, 2, 2)
c.fill = yellow_fill; c.border = thick_border; c.font = Font(name='Arial', bold=True, size=12)

ws3.cell(7, 1, 'Outside Rep Comp:').font = bold_font
c = sc(ws3, 7, 2, 100000, money_fmt)
c.fill = yellow_fill; c.border = thick_border

ws3.cell(8, 1, 'Inside Rep Comp:').font = bold_font
c = sc(ws3, 8, 2, 65000, money_fmt)
c.fill = yellow_fill; c.border = thick_border

# Revenue assumptions
ws3.cell(10, 1, 'REVENUE PER REP (historical averages — edit to adjust)').font = section_font
ws3.merge_cells('A10:F10')
for c, h in enumerate(['', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'], 1):
    ws3.cell(11, c, h)
hdr(ws3, 11, 6)

ws3.cell(12, 1, 'Outside Rep Revenue').font = bold_font
outside_rev = [300743, 375408, 528493, 800751, 890622]
for c, v in enumerate(outside_rev, 2):
    cell = sc(ws3, 12, c, v, money_fmt)
    cell.fill = yellow_fill

ws3.cell(13, 1, 'Inside Rep Revenue').font = bold_font
inside_rev = [194837, 223726, 279838, 292239, 408323]
for c, v in enumerate(inside_rev, 2):
    cell = sc(ws3, 13, c, v, money_fmt)
    cell.fill = yellow_fill

# Buildup table
ws3.cell(15, 1, 'REVENUE BUILDUP').font = section_font
for c, h in enumerate(['', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5', '5-Yr Total'], 1):
    ws3.cell(16, c, h)
hdr(ws3, 16, 7)

# Outside reps rows
for i in range(5):
    row = 17 + i
    sc(ws3, row, 1, f'Outside Rep {i+1}', bold=True)
    for yr_col in range(2, 7):
        ref = f'{get_column_letter(yr_col)}$12'
        sc(ws3, row, yr_col, f'=IF({i+1}<=$B$5,{ref},0)', money_fmt)
    sc(ws3, row, 7, f'=SUM(B{row}:F{row})', money_fmt)
    if i % 2 == 0:
        for c in range(1, 8): ws3.cell(row, c).fill = green_fill

# Inside reps rows
for i in range(5):
    row = 22 + i
    sc(ws3, row, 1, f'Inside Rep {i+1}', bold=True)
    for yr_col in range(2, 7):
        ref = f'{get_column_letter(yr_col)}$13'
        sc(ws3, row, yr_col, f'=IF({i+1}<=$B$6,{ref},0)', money_fmt)
    sc(ws3, row, 7, f'=SUM(B{row}:F{row})', money_fmt)
    if i % 2 == 0:
        for c in range(1, 8): ws3.cell(row, c).fill = light_fill

# Totals
tr = 28
for c, h in enumerate(['', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5', '5-Yr Total'], 1):
    ws3.cell(tr, c, h)
hdr(ws3, tr, 7)

sc(ws3, tr+1, 1, 'Outside Total', bold=True)
for col in range(2, 8):
    cl = get_column_letter(col)
    sc(ws3, tr+1, col, f'=SUM({cl}17:{cl}21)', money_fmt)
for c in range(1, 8): ws3.cell(tr+1, c).fill = green_fill

sc(ws3, tr+2, 1, 'Inside Total', bold=True)
for col in range(2, 8):
    cl = get_column_letter(col)
    sc(ws3, tr+2, col, f'=SUM({cl}22:{cl}26)', money_fmt)
for c in range(1, 8): ws3.cell(tr+2, c).fill = light_fill

sc(ws3, tr+3, 1, 'COMBINED TOTAL', bold=True)
for col in range(2, 8):
    cl = get_column_letter(col)
    cell = sc(ws3, tr+3, col, f'={cl}{tr+1}+{cl}{tr+2}', money_fmt)
    cell.font = total_font; cell.fill = total_fill
ws3.cell(tr+3, 1).font = total_font; ws3.cell(tr+3, 1).fill = total_fill; ws3.cell(tr+3, 1).border = thin_border

# Investment analysis
ir = tr + 5
ws3.cell(ir, 1, 'INVESTMENT ANALYSIS').font = section_font
for c, h in enumerate(['', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'], 1):
    ws3.cell(ir+1, c, h)
hdr(ws3, ir+1, 6)

sc(ws3, ir+2, 1, 'Total Comp Cost', bold=True)
for col in range(2, 7):
    sc(ws3, ir+2, col, f'=$B$5*$B$7+$B$6*$B$8', money_fmt)

sc(ws3, ir+3, 1, 'Total Revenue', bold=True)
for col in range(2, 7):
    cl = get_column_letter(col)
    sc(ws3, ir+3, col, f'={cl}{tr+3}', money_fmt)

sc(ws3, ir+4, 1, 'Net (Rev - Comp)', bold=True)
for col in range(2, 7):
    cl = get_column_letter(col)
    cell = sc(ws3, ir+4, col, f'={cl}{ir+3}-{cl}{ir+2}', money_fmt)
    cell.font = Font(name='Arial', bold=True, size=11)

sc(ws3, ir+5, 1, 'ROI', bold=True)
for col in range(2, 7):
    cl = get_column_letter(col)
    sc(ws3, ir+5, col, f'=IF({cl}{ir+2}>0,{cl}{ir+4}/{cl}{ir+2},0)', pct_fmt)

set_widths(ws3, [28, 14, 14, 14, 14, 14, 16])

# ============================================================
# SHEET 4: Territory Case Studies
# ============================================================
ws4 = wb.create_sheet('Lost Territories')
ws4.sheet_properties.tabColor = 'FF0000'

ws4['A1'] = 'Revenue Lost When Territory Reps Leave'
ws4['A1'].font = title_font
ws4.merge_cells('A1:G1')

for c, h in enumerate(['Rep', 'Territory', 'Years Active', 'Peak Annual Rev', 'Career Revenue', 'Clients', 'Left'], 1):
    ws4.cell(3, c, h)
hdr(ws4, 3, 7)

lost = [
    ('Lindsey Davis', 'MD/DC', '5yr', 1629390, 6548594, 19, 2010),
    ('Tom Kane', 'PA', '8yr', 1137678, 4469457, 89, 2023),
    ('Eric Schwarzel', 'PA/NJ/DE', '3yr', 784957, 1263399, 38, 2010),
    ('Courtney Blue', 'OR/TX', '3yr', 478997, 1250695, 20, 2010),
    ('Danielle Penge', 'OR/NJ', '3yr', 386599, 877655, 27, 2010),
    ('Michael Woessner', 'NJ/OR', '4yr', 452289, 885795, 23, 2009),
    ('Kara Klund', 'IL/TX', '2yr', 562615, 906112, 20, 2008),
    ('Melissa Fox', 'MD/DC/VA', '2yr', 281443, 407129, 9, 2018),
]
for i, (name, terr, yrs, peak, career, clients, left) in enumerate(lost):
    row = 4 + i
    sc(ws4, row, 1, name, bold=True)
    sc(ws4, row, 2, terr)
    sc(ws4, row, 3, yrs)
    sc(ws4, row, 4, peak, money_fmt)
    sc(ws4, row, 5, career, money_fmt)
    sc(ws4, row, 6, clients)
    sc(ws4, row, 7, left)
    ws4.cell(row, 7).number_format = '0'
    if i % 2 == 0:
        for c in range(1, 8): ws4.cell(row, c).fill = light_fill

ws4.cell(13, 1, 'TOTAL LOST REVENUE:').font = Font(name='Arial', bold=True, size=12, color='FF0000')
sc(ws4, 13, 2, f'=SUM(E4:E11)', money_fmt).font = Font(name='Arial', bold=True, size=12, color='FF0000')

ws4.cell(15, 1, 'Ben Taub Case Study (Courtney Blue\'s territory):').font = section_font
ws4.merge_cells('A15:G15')
for c, h in enumerate(['Year', 'Rep', 'Invoices', 'Revenue', 'Status'], 1):
    ws4.cell(16, c, h)
hdr(ws4, 16, 5)
bentaub = [
    (2006, 'Molly Kohn', 222, 129595, 'Active'),
    (2007, 'Molly Kohn / Courtney Blue', 321, 162282, 'Transition'),
    (2008, 'Courtney Blue', 255, 139136, 'Active'),
    (2009, 'Courtney Blue', 155, 111401, 'Active'),
    (2010, 'Courtney Blue (leaving)', 15, 13235, 'Winding down'),
    (2011, 'No rep', 0, 0, 'DEAD'),
]
for i, (yr, rep, inv, rev, status) in enumerate(bentaub):
    row = 17 + i
    sc(ws4, row, 1, yr); ws4.cell(row, 1).number_format = '0'
    sc(ws4, row, 2, rep)
    sc(ws4, row, 3, inv)
    sc(ws4, row, 4, rev, money_fmt)
    sc(ws4, row, 5, status)
    if status == 'DEAD':
        for c in range(1, 6): ws4.cell(row, c).fill = red_fill

set_widths(ws4, [28, 14, 12, 16, 16, 10, 10])

# ============================================================
# SHEET 5: Company Context
# ============================================================
ws5 = wb.create_sheet('Company Totals')
ws5.sheet_properties.tabColor = 'FFC000'

ws5['A1'] = 'TSI Company Revenue Context'
ws5['A1'].font = title_font
for c, h in enumerate(['Year', 'Total Repairs', 'Revenue', 'Clients', 'Avg Ticket'], 1):
    ws5.cell(3, c, h)
hdr(ws5, 3, 5)
company = [
    (2022, 6701, 5507495, 539, 822),
    (2023, 6911, 4948301, 494, 716),
    (2024, 7626, 7122944, 624, 934),
    (2025, 9915, 9059428, 727, 914),
    (2026, 1591, 1559126, 324, 980),
]
for i, (yr, repairs, rev, clients, avg) in enumerate(company):
    row = 4 + i
    sc(ws5, row, 1, yr); ws5.cell(row, 1).number_format = '0'
    sc(ws5, row, 2, repairs, '#,##0')
    sc(ws5, row, 3, rev, money_fmt)
    sc(ws5, row, 4, clients, '#,##0')
    sc(ws5, row, 5, avg, money_fmt)
ws5.cell(9, 1, '* 2026 data through March 13 only').font = Font(name='Arial', italic=True, size=9, color='FF0000')
set_widths(ws5, [10, 14, 16, 12, 12])

# ============================================================
# Save
# ============================================================
out = r'C:\tmp\tsi-sales-rep-benchmarks-v3.xlsx'
wb.save(out)
print(f'Saved to {out}')
print('Sheets:', [s.title for s in wb.worksheets])
