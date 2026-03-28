import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

hdr_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='143B87')
normal = Font(name='Arial', size=10)
bold = Font(name='Arial', bold=True, size=10)
section = Font(name='Arial', bold=True, size=12, color='143B87')
thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
thick = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
money = '$#,##0;($#,##0);"-"'
pct = '0.0%'
yellow = PatternFill('solid', fgColor='FFFF00')
green = PatternFill('solid', fgColor='C6EFCE')
light = PatternFill('solid', fgColor='D6E4F0')
orange = PatternFill('solid', fgColor='FCE4D6')
navy = PatternFill('solid', fgColor='143B87')
navy_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
red_font = Font(name='Arial', bold=True, size=16, color='FF0000')
big_font = Font(name='Arial', bold=True, size=14, color='143B87')

def hdr(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True); cell.border = thin

def sc(ws, r, c, val=None, fmt=None, b=False):
    cell = ws.cell(r, c)
    if val is not None: cell.value = val
    cell.font = bold if b else normal; cell.border = thin
    if fmt: cell.number_format = fmt
    return cell

def navy_row(ws, r, cols):
    for c in range(1, cols + 1):
        ws.cell(r, c).font = navy_font; ws.cell(r, c).fill = navy; ws.cell(r, c).border = thin

# ============================================================
# SHEET 1: The Planner
# ============================================================
ws = wb.active
ws.title = 'Revenue Planner'
ws.sheet_properties.tabColor = '143B87'

ws['A1'] = 'TSI Revenue Planner'
ws['A1'].font = Font(name='Arial', bold=True, size=20, color='143B87')
ws.merge_cells('A1:E1')

# ─── GOAL ───
r = 3
sc(ws, r, 1, 'Revenue Goal', b=True).font = Font(name='Arial', bold=True, size=14)
c = sc(ws, r, 2, 20000000, money)
c.fill = yellow; c.border = thick; c.font = Font(name='Arial', bold=True, size=18, color='FF0000')

# ─── CURRENT ROSTER ───
r = 5
ws.cell(r, 1, 'WHAT WE HAVE NOW').font = section
r += 1
for c, h in enumerate(['Rep', 'Type', 'Annual Revenue'], 1): ws.cell(r, c, h)
hdr(ws, r, 3)
roster_hdr = r
r += 1

# Current reps: avg 2024+2025
current = [
    ('Brian Kenney', 'Inside', 5134556),
    ('Debbie Hightower', 'Inside', 1627084),
    ('Bernie DeLacy', 'Outside', 1750205),
    ('Brandi Cook', 'Inside', 1815826),
    ('Seamus Glavin', 'Inside', 1139552),
    ('Jim Rygiel', 'Outside', 442609),
    ('Timothy Reilly', 'Inside', 487810),
    ('Donald Haynes', 'Inside', 125999),
    ('Ryan George', 'Inside', 238269),
    ('BSC Programs (combined)', 'Program', 2198651),
]

roster_first = r
for name, rtype, rev in current:
    sc(ws, r, 1, name, b=True)
    sc(ws, r, 2, rtype)
    c = sc(ws, r, 3, rev, money)
    c.fill = yellow  # editable
    fill = green if rtype == 'Outside' else orange if rtype == 'Program' else light
    ws.cell(r, 1).fill = fill; ws.cell(r, 2).fill = fill
    r += 1
roster_last = r - 1

# Current total
sc(ws, r, 1, 'CURRENT TOTAL', b=True)
sc(ws, r, 3, f'=SUM(C{roster_first}:C{roster_last})', money)
navy_row(ws, r, 3)
current_total = r
r += 1

# Gap
sc(ws, r, 1, 'GAP TO GOAL', b=True)
sc(ws, r, 3, f'=B3-C{current_total}', money).font = red_font
ws.cell(r, 1).font = Font(name='Arial', bold=True, size=14, color='FF0000')
gap_row = r
r += 2

# ─── ADD REPS ───
ws.cell(r, 1, 'ADD REPS TO CLOSE THE GAP').font = section
r += 1
ws.cell(r, 1, 'Pick Inside or Outside. Revenue auto-fills from historical avg. Yellow = editable.').font = Font(name='Arial', italic=True, size=9, color='666666')
ws.merge_cells(f'A{r}:E{r}')
r += 1

for c, h in enumerate(['New Rep', 'Type', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'], 1):
    ws.cell(r, c, h)
hdr(ws, r, 7)
hire_hdr = r
r += 1

# Assumptions (hidden reference area)
# Outside yr1-5: 348K, 453K, 491K, 815K, 1456K
# Inside yr1-5: 202K, 219K, 301K, 313K, 391K
# Put these in a reference area
ref_row = 60
sc(ws, ref_row, 1, 'Outside Yr1-5 Revenue Benchmarks (do not edit)')
outside_bench = [348374, 452886, 491083, 814695, 1455687]
inside_bench = [202000, 218961, 301268, 312861, 390956]
for i, v in enumerate(outside_bench):
    sc(ws, ref_row + 1, i + 1, v, money)  # row 61, cols A-E
for i, v in enumerate(inside_bench):
    sc(ws, ref_row + 2, i + 1, v, money)  # row 62, cols A-E
sc(ws, ref_row + 2, 1, inside_bench[0], money)

# 10 new rep rows
hire_first = r
for i in range(10):
    rep_num = i + 1
    sc(ws, r, 1, f'New Rep {rep_num}')
    # Type - default alternating
    c = sc(ws, r, 2, 'Outside' if i < 3 else 'Inside' if i < 7 else '')
    c.fill = yellow
    # Year 1-5: lookup based on type
    for yr in range(1, 6):
        col = yr + 2
        # If type is blank, show 0. If Outside, use row 61. If Inside, use row 62.
        ref_col = get_column_letter(yr)
        formula = f'=IF(B{r}="Outside",{ref_col}{ref_row+1},IF(B{r}="Inside",{ref_col}{ref_row+2},0))'
        sc(ws, r, col, formula, money)
    # Alternating fill
    if i % 2 == 0:
        for c in range(1, 8): ws.cell(r, c).fill = PatternFill('solid', fgColor='F2F2F2')
    r += 1
hire_last = r - 1

# New hire totals per year
sc(ws, r, 1, 'NEW HIRES TOTAL', b=True)
for col in range(3, 8):
    cl = get_column_letter(col)
    sc(ws, r, col, f'=SUM({cl}{hire_first}:{cl}{hire_last})', money)
navy_row(ws, r, 7)
hire_total = r
r += 2

# ─── THE ANSWER ───
ws.cell(r, 1, 'THE ANSWER').font = Font(name='Arial', bold=True, size=16, color='143B87')
r += 1

for c, h in enumerate(['', '', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'], 1):
    ws.cell(r, c, h)
hdr(ws, r, 7)
r += 1

sc(ws, r, 1, 'Current Roster', b=True)
for col in range(3, 8): sc(ws, r, col, f'=C{current_total}', money)
for c in range(1, 8): ws.cell(r, c).fill = light
r += 1

sc(ws, r, 1, 'New Hire Revenue', b=True)
for col in range(3, 8):
    cl = get_column_letter(col)
    sc(ws, r, col, f'={cl}{hire_total}', money)
for c in range(1, 8): ws.cell(r, c).fill = green
r += 1

sc(ws, r, 1, 'TOTAL REVENUE', b=True)
for col in range(3, 8):
    cl = get_column_letter(col)
    sc(ws, r, col, f'={cl}{r-2}+{cl}{r-1}', money)
navy_row(ws, r, 7)
proj_total = r
r += 1

sc(ws, r, 1, 'Goal', b=True)
for col in range(3, 8): sc(ws, r, col, '=$B$3', money)
r += 1

sc(ws, r, 1, 'OVER / (SHORT)', b=True)
ws.cell(r, 1).font = red_font
for col in range(3, 8):
    cl = get_column_letter(col)
    cell = sc(ws, r, col, f'={cl}{proj_total}-{cl}{r-1}', money)
    cell.font = Font(name='Arial', bold=True, size=14, color='FF0000')

# Column widths
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 12
for c in range(3, 8): ws.column_dimensions[get_column_letter(c)].width = 16

# ============================================================
# SHEET 2: Inside vs Outside (reference)
# ============================================================
ws2 = wb.create_sheet('Inside vs Outside')
ws2.sheet_properties.tabColor = '143B87'
ws2['A1'] = 'What the Data Says: Inside vs Outside Reps'
ws2['A1'].font = Font(name='Arial', bold=True, size=16, color='143B87')
ws2.merge_cells('A1:F1')
ws2['A2'] = '25 years | 200K+ invoices | 8 outside territory reps | 100+ inside reps'
ws2['A2'].font = Font(name='Arial', italic=True, size=10, color='666666')

r = 4
ws2.cell(r, 1, 'AVERAGE REP REVENUE BY YEAR').font = section
r += 1
for c, h in enumerate(['', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'], 1): ws2.cell(r, c, h)
hdr(ws2, r, 6)

for label, vals, fill in [
    ('Outside Territory Rep', [348374, 452886, 491083, 814695, 1455687], green),
    ('Inside / Phone Rep', [202000, 218961, 301268, 312861, 390956], light),
]:
    r += 1
    sc(ws2, r, 1, label, b=True)
    for c, v in enumerate(vals, 2): sc(ws2, r, c, v, money)
    for c in range(1, 7): ws2.cell(r, c).fill = fill

r += 2
ws2.cell(r, 1, 'KEY FACTS').font = section
facts = [
    'Outside reps: 80% higher revenue Year 1 — but they need an inherited book to start.',
    'Inside reps: land contracts by Year 2 ($44K avg). Outside reps = 100% FFS.',
    'Inside reps source more net new clients early. Outside reps find bigger accounts later.',
    'Only 19% of inside reps make it to 5 years. 53% gone by Year 3.',
    'Contract revenue = 36% of company total. Contracts come from inside relationships.',
    'Outside reps who stay hit $800K+ by Year 4. But only 8 ever did it at TSI.',
    'Ben Taub: $140K/yr when Courtney Blue was there. $0 after she left.',
]
for i, f in enumerate(facts):
    r += 1
    ws2.cell(r, 1, f'  {i+1}. {f}').font = normal
    ws2.merge_cells(f'A{r}:F{r}')

for c in range(1, 7): ws2.column_dimensions[get_column_letter(c)].width = 16
ws2.column_dimensions['A'].width = 34

# Save
out = r'C:\tmp\tsi-sales-rep-planner-v7.xlsx'
wb.save(out)
print(f'Saved to {out}')
