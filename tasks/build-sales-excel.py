import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.label import DataLabelList

wb = openpyxl.Workbook()

# Styles
hdr_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='143B87')
sub_hdr_fill = PatternFill('solid', fgColor='4472C4')
light_fill = PatternFill('solid', fgColor='D6E4F0')
green_fill = PatternFill('solid', fgColor='C6EFCE')
red_fill = PatternFill('solid', fgColor='FFC7CE')
yellow_fill = PatternFill('solid', fgColor='FFEB9C')
title_font = Font(name='Arial', bold=True, size=14, color='143B87')
section_font = Font(name='Arial', bold=True, size=12, color='143B87')
normal_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', bold=True, size=10)
money_fmt = '$#,##0;($#,##0);"-"'
pct_fmt = '0.0%'
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def style_data_cell(ws, row, col, fmt=None):
    cell = ws.cell(row=row, column=col)
    cell.font = normal_font
    cell.border = thin_border
    if fmt:
        cell.number_format = fmt
    return cell

# CPI multipliers to adjust to 2026 dollars
# Source: BLS CPI-U annual averages, base 2026 est ~320
cpi = {
    1994: 148.2, 1995: 152.4, 1996: 156.9, 1997: 160.5, 1998: 163.0,
    1999: 166.6, 2000: 172.2, 2001: 177.1, 2002: 179.9, 2003: 184.0,
    2004: 188.9, 2005: 195.3, 2006: 201.6, 2007: 207.3, 2008: 215.3,
    2009: 214.5, 2010: 218.1, 2011: 224.9, 2012: 229.6, 2013: 233.0,
    2014: 236.7, 2015: 237.0, 2016: 240.0, 2017: 245.1, 2018: 251.1,
    2019: 255.7, 2020: 258.8, 2021: 271.0, 2022: 292.7, 2023: 304.7,
    2024: 314.2, 2025: 320.0, 2026: 325.0
}
cpi_2026 = cpi[2026]

# ============================================================
# SHEET 1: Executive Summary
# ============================================================
ws = wb.active
ws.title = 'Executive Summary'
ws.sheet_properties.tabColor = '143B87'

ws['A1'] = 'TSI Sales Rep Performance Benchmarks'
ws['A1'].font = Font(name='Arial', bold=True, size=18, color='143B87')
ws.merge_cells('A1:G1')
ws['A2'] = 'Data Source: WinScopeNet Invoice Table | 109 Real Reps | 25 Years (1994-2026) | 200K+ Invoices'
ws['A2'].font = Font(name='Arial', italic=True, size=10, color='666666')
ws.merge_cells('A2:G2')

ws['A4'] = 'KEY FINDING: Net New Revenue Crossover'
ws['A4'].font = section_font
ws['A5'] = 'A rep\'s self-sourced (net new) revenue overtakes inherited revenue in Year 5.'
ws['A5'].font = normal_font
ws['A6'] = 'Cold territory with NO inherited book = ~$44K Year 1 revenue.'
ws['A6'].font = Font(name='Arial', bold=True, size=10, color='FF0000')

r = 8
ws.cell(r, 1, 'Tenure Year').font = bold_font
ws.cell(r, 2, 'Inherited Revenue').font = bold_font
ws.cell(r, 3, 'Net New Revenue').font = bold_font
ws.cell(r, 4, 'Combined').font = bold_font
ws.cell(r, 5, 'Net New %').font = bold_font
ws.cell(r, 6, '# Reps (Inherited)').font = bold_font
ws.cell(r, 7, '# Reps (Net New)').font = bold_font
style_header_row(ws, r, 7)

ramp_data = [
    (1, 187829, 44085, 95, 101),
    (2, 200451, 74028, 74, 77),
    (3, 194140, 146008, 45, 46),
    (4, 200941, 176736, 32, 33),
    (5, 235074, 257577, 22, 22),
    (6, 277232, 313550, 15, 15),
    (7, 318061, 290813, 13, 14),
    (8, 296330, 301286, 12, 11),
    (9, 381545, 247738, 8, 8),
    (10, 372818, 201986, 6, 8),
]

for i, (yr, inh, nn, reps_i, reps_n) in enumerate(ramp_data):
    row = r + 1 + i
    style_data_cell(ws, row, 1).value = f'Year {yr}'
    style_data_cell(ws, row, 2, money_fmt).value = inh
    style_data_cell(ws, row, 3, money_fmt).value = nn
    style_data_cell(ws, row, 4, money_fmt).value = f'=B{row}+C{row}'
    style_data_cell(ws, row, 5, pct_fmt).value = f'=C{row}/D{row}'
    style_data_cell(ws, row, 6).value = reps_i
    style_data_cell(ws, row, 7).value = reps_n
    if yr == 5:
        for c in range(1, 8):
            ws.cell(row, c).fill = green_fill

r2 = r + 12
ws.cell(r2, 1, 'What This Means for a New Territory').font = section_font
ws.cell(r2+1, 1, 'Scenario').font = bold_font
ws.cell(r2+1, 2, 'Year 1').font = bold_font
ws.cell(r2+1, 3, 'Year 2').font = bold_font
ws.cell(r2+1, 4, 'Year 3').font = bold_font
ws.cell(r2+1, 5, 'Year 5').font = bold_font
style_header_row(ws, r2+1, 5)

ws.cell(r2+2, 1, 'Cold Territory (no book)').font = bold_font
ws.cell(r2+2, 1).fill = red_fill
for c, v in [(2, 44085), (3, 74028), (4, 146008), (5, 257577)]:
    style_data_cell(ws, r2+2, c, money_fmt).value = v

ws.cell(r2+3, 1, 'With Inherited Book').font = bold_font
ws.cell(r2+3, 1).fill = green_fill
for c, v in [(2, 231914), (3, 274479), (4, 340148), (5, 492651)]:
    style_data_cell(ws, r2+3, c, money_fmt).value = v

for col in range(1, 8):
    ws.column_dimensions[get_column_letter(col)].width = 18

# Add chart
chart = BarChart()
chart.type = 'col'
chart.grouping = 'stacked'
chart.title = 'Average Rep Revenue by Tenure Year: Net New vs Inherited'
chart.y_axis.title = 'Revenue ($)'
chart.x_axis.title = 'Tenure Year'
chart.style = 10
cats = Reference(ws, min_col=1, min_row=9, max_row=18)
inherited = Reference(ws, min_col=2, min_row=8, max_row=18)
netnew = Reference(ws, min_col=3, min_row=8, max_row=18)
chart.add_data(inherited, titles_from_data=True)
chart.add_data(netnew, titles_from_data=True)
chart.set_categories(cats)
chart.series[0].graphicalProperties.solidFill = '4472C4'
chart.series[1].graphicalProperties.solidFill = '70AD47'
chart.width = 20
chart.height = 12
ws.add_chart(chart, f'A{r2+6}')

# ============================================================
# SHEET 2: Ramp Curve Detail
# ============================================================
ws2 = wb.create_sheet('Rep Ramp Curve')
ws2.sheet_properties.tabColor = '4472C4'

ws2['A1'] = 'Average Sales Rep Revenue Ramp by Tenure Year'
ws2['A1'].font = title_font
ws2.merge_cells('A1:H1')
ws2['A2'] = 'Based on 109 real reps across 25 years of invoice data (1994-2026)'
ws2['A2'].font = Font(name='Arial', italic=True, size=10, color='666666')

headers = ['Tenure Year', 'Inherited ($)', 'Net New ($)', 'Combined ($)', 'Net New %',
           'Inherited Clients', 'Net New Clients', 'Total Clients']
for c, h in enumerate(headers, 1):
    ws2.cell(4, c, h)
style_header_row(ws2, 4, 8)

inh_clients = [15, 14, 13, 13, 16, 21, 17, 18, 20, 24]
nn_clients = [9, 10, 12, 14, 18, 26, 24, 26, 27, 23]

for i, (yr, inh, nn, reps_i, reps_n) in enumerate(ramp_data):
    row = 5 + i
    style_data_cell(ws2, row, 1).value = f'Year {yr}'
    style_data_cell(ws2, row, 2, money_fmt).value = inh
    style_data_cell(ws2, row, 3, money_fmt).value = nn
    style_data_cell(ws2, row, 4, money_fmt).value = f'=B{row}+C{row}'
    style_data_cell(ws2, row, 5, pct_fmt).value = f'=C{row}/D{row}'
    style_data_cell(ws2, row, 6).value = inh_clients[i]
    style_data_cell(ws2, row, 7).value = nn_clients[i]
    style_data_cell(ws2, row, 8).value = f'=F{row}+G{row}'

for col in range(1, 9):
    ws2.column_dimensions[get_column_letter(col)].width = 16

# Inflation adjustment section
ws2.cell(17, 1, 'INFLATION ADJUSTED (2026 Dollars)').font = section_font
ws2.merge_cells('A17:H17')
ws2['A18'] = 'Note: Uses CPI-U multipliers. Older reps\' revenue adjusted upward to reflect today\'s purchasing power.'
ws2['A18'].font = Font(name='Arial', italic=True, size=9, color='666666')
ws2.merge_cells('A18:H18')

headers2 = ['Start Year (approx)', 'CPI at Start', 'CPI 2026 (est)', 'Multiplier',
            'Yr1 Nominal', 'Yr1 Adjusted', 'Yr3 Nominal', 'Yr3 Adjusted']
for c, h in enumerate(headers2, 1):
    ws2.cell(20, c, h)
style_header_row(ws2, 20, 8)

sample_years = [2000, 2005, 2010, 2015, 2020, 2024]
for i, sy in enumerate(sample_years):
    row = 21 + i
    mult = cpi_2026 / cpi[sy]
    style_data_cell(ws2, row, 1).value = sy
    style_data_cell(ws2, row, 1).number_format = '0'
    style_data_cell(ws2, row, 2, '0.0').value = cpi[sy]
    style_data_cell(ws2, row, 3, '0.0').value = cpi_2026
    style_data_cell(ws2, row, 4, '0.00x').value = mult
    style_data_cell(ws2, row, 5, money_fmt).value = 44085
    style_data_cell(ws2, row, 6, money_fmt).value = round(44085 * mult)
    style_data_cell(ws2, row, 7, money_fmt).value = 146008
    style_data_cell(ws2, row, 8, money_fmt).value = round(146008 * mult)

# ============================================================
# SHEET 3: Top Rep Profiles
# ============================================================
ws3 = wb.create_sheet('Top Rep Profiles')
ws3.sheet_properties.tabColor = '70AD47'

ws3['A1'] = 'Individual Rep Performance Profiles'
ws3['A1'].font = title_font
ws3.merge_cells('A1:I1')

# Data: rep name, start year, [(yr, inherited, net_new, inh_clients, nn_clients)]
top_reps = [
    ('Brian Kenney', 2011, [
        (1, 79167, 52561, 14, 11), (2, 232762, 128068, 24, 17),
        (3, 338999, 547687, 22, 34), (4, 456044, 572622, 21, 38),
        (5, 606843, 802001, 23, 46), (6, 768359, 902021, 29, 45),
        (7, 867480, 883907, 29, 53), (8, 845223, 828220, 37, 63),
        (9, 532653, 503738, 31, 60), (10, 520734, 364092, 21, 40),
    ]),
    ('Debbie Hightower', 2024, [
        (1, 601945, 282384, 58, 58), (2, 882780, 573448, 50, 75),
        (3, 254840, 266766, 18, 39), (4, 299048, 285364, 20, 42),
        (5, 294011, 264796, 21, 37), (6, 319882, 560708, 25, 47),
        (7, 484032, 534885, 24, 51), (8, 456613, 679368, 22, 45),
        (9, 565809, 535110, 23, 45), (10, 399056, 646652, 24, 48),
    ]),
    ('Gina Ellis', 2008, [
        (1, 342151, 92169, 15, 19), (2, 894311, 135111, 23, 27),
        (3, 952484, 178115, 31, 30), (4, 791071, 324381, 29, 26),
        (5, 721367, 365356, 22, 24), (6, 755749, 387118, 25, 31),
        (7, 733952, 353523, 21, 25), (8, 187325, 129328, 16, 16),
    ]),
    ('Jim Rygiel', 2024, [
        (1, 172697, 263293, 1, 9), (2, 274625, 155032, 3, 11),
    ]),
    ('Courtney Blue', 2007, [
        (1, 396039, 48116, 10, 10), (2, 420377, 58620, 6, 5),
        (3, 272413, 55130, 6, 5),
    ]),
    ('Fausto Guillermo', 2021, [
        (1, 158365, 109016, 24, 7), (2, 764414, 259269, 39, 17),
        (3, 656997, 247461, 35, 21), (4, 262262, 63773, 24, 14),
    ]),
    ('Ryan Gades', 2007, [
        (1, 103944, 0, 19, 0), (2, 192449, 0, 41, 0),
        (3, 234499, 0, 40, 0), (4, 525413, 0, 64, 0),
        (5, 925976, 0, 93, 0), (6, 1179745, 0, 88, 0),
        (7, 1147619, 0, 76, 0),
    ]),
    ('John Luedke', 2005, [
        (1, 388948, 59069, 35, 21), (2, 449058, 252016, 32, 41),
        (3, 361558, 377542, 28, 37), (4, 344785, 345221, 20, 53),
        (5, 264934, 402482, 20, 40), (6, 244209, 317829, 22, 40),
        (7, 325220, 307349, 19, 36), (8, 501630, 516453, 18, 42),
        (9, 693951, 296066, 20, 38), (10, 271688, 235222, 24, 38),
    ]),
    ('Mark DeBiase', 2024, [
        (1, 0, 437098, 0, 60), (2, 0, 996292, 0, 92),
    ]),
    ('Bill Odonnell', 2010, [
        (1, 60428, 43516, 7, 12), (2, 81417, 111032, 14, 27),
        (3, 22757, 211742, 10, 30), (4, 140867, 384546, 16, 48),
        (5, 109793, 816184, 17, 76), (6, 82228, 1097517, 11, 77),
        (7, 132895, 1014724, 9, 67), (8, 117229, 640220, 8, 39),
    ]),
]

current_row = 3
for rep_name, start_yr, years in top_reps:
    ws3.cell(current_row, 1, rep_name).font = section_font
    ws3.cell(current_row, 2, f'Started: {start_yr}').font = Font(name='Arial', italic=True, size=10, color='666666')
    current_row += 1

    for c, h in enumerate(['Year', 'Inherited', 'Net New', 'Combined', 'Net New %', 'Inh Clients', 'NN Clients', 'Total Clients'], 1):
        ws3.cell(current_row, c, h)
    style_header_row(ws3, current_row, 8)
    current_row += 1

    for yr, inh, nn, ic, nc in years:
        style_data_cell(ws3, current_row, 1).value = f'Year {yr}'
        style_data_cell(ws3, current_row, 2, money_fmt).value = inh
        style_data_cell(ws3, current_row, 3, money_fmt).value = nn
        style_data_cell(ws3, current_row, 4, money_fmt).value = inh + nn
        if inh + nn > 0:
            style_data_cell(ws3, current_row, 5, pct_fmt).value = nn / (inh + nn)
        else:
            style_data_cell(ws3, current_row, 5).value = '-'
        style_data_cell(ws3, current_row, 6).value = ic
        style_data_cell(ws3, current_row, 7).value = nc
        style_data_cell(ws3, current_row, 8).value = ic + nc
        current_row += 1

    current_row += 1

for col in range(1, 9):
    ws3.column_dimensions[get_column_letter(col)].width = 15

# ============================================================
# SHEET 4: Company Totals
# ============================================================
ws4 = wb.create_sheet('Company Totals')
ws4.sheet_properties.tabColor = 'FFC000'

ws4['A1'] = 'TSI Company Revenue Context'
ws4['A1'].font = title_font
ws4.merge_cells('A1:F1')

for c, h in enumerate(['Year', 'Total Repairs', 'Total Revenue', 'Active Clients', 'Avg Ticket', 'YoY Growth'], 1):
    ws4.cell(3, c, h)
style_header_row(ws4, 3, 6)

company = [
    (2022, 6701, 5507495, 539, 822),
    (2023, 6911, 4948301, 494, 716),
    (2024, 7626, 7122944, 624, 934),
    (2025, 9915, 9059428, 727, 914),
    (2026, 1591, 1559126, 324, 980),
]
for i, (yr, repairs, rev, clients, avg) in enumerate(company):
    row = 4 + i
    style_data_cell(ws4, row, 1).value = yr
    ws4.cell(row, 1).number_format = '0'
    style_data_cell(ws4, row, 2, '#,##0').value = repairs
    style_data_cell(ws4, row, 3, money_fmt).value = rev
    style_data_cell(ws4, row, 4, '#,##0').value = clients
    style_data_cell(ws4, row, 5, money_fmt).value = avg
    if i > 0:
        style_data_cell(ws4, row, 6, pct_fmt).value = f'=C{row}/C{row-1}-1'

ws4['A10'] = '* 2026 data is partial (through March 13, 2026)'
ws4['A10'].font = Font(name='Arial', italic=True, size=9, color='FF0000')

for col in range(1, 7):
    ws4.column_dimensions[get_column_letter(col)].width = 16

# ============================================================
# SHEET 5: Territory Case Studies
# ============================================================
ws5 = wb.create_sheet('Territory Case Studies')
ws5.sheet_properties.tabColor = 'FF0000'

ws5['A1'] = 'What Happens When a Territory Loses Its Rep'
ws5['A1'].font = title_font
ws5.merge_cells('A1:F1')

ws5['A3'] = 'Ben Taub General Hospital (Houston, TX)'
ws5['A3'].font = section_font
ws5['A4'] = 'Bill-to: Philips Medical Systems | Currently assigned to John Luedke (inactive)'
ws5['A4'].font = Font(name='Arial', italic=True, size=10)

for c, h in enumerate(['Year', 'Rep', 'Invoices', 'Revenue', 'Status'], 1):
    ws5.cell(6, c, h)
style_header_row(ws5, 6, 5)

bentaub = [
    (2006, 'Molly Kohn', 222, 129595, 'Active'),
    (2007, 'Molly Kohn / Courtney Blue', 321, 162282, 'Transition'),
    (2008, 'Courtney Blue', 255, 139136, 'Active'),
    (2009, 'Courtney Blue', 155, 111401, 'Active'),
    (2010, 'Courtney Blue → Gina Ellis', 15, 13235, 'Rep left'),
    (2011, '(no rep)', 0, 0, 'Dead'),
]
for i, (yr, rep, inv, rev, status) in enumerate(bentaub):
    row = 7 + i
    style_data_cell(ws5, row, 1).value = yr
    ws5.cell(row, 1).number_format = '0'
    style_data_cell(ws5, row, 2).value = rep
    style_data_cell(ws5, row, 3, '#,##0').value = inv
    style_data_cell(ws5, row, 4, money_fmt).value = rev
    style_data_cell(ws5, row, 5).value = status
    if status == 'Dead':
        for c in range(1, 6):
            ws5.cell(row, c).fill = red_fill

ws5['A15'] = 'Lost Territory Revenue — Reps Who Left'
ws5['A15'].font = section_font

for c, h in enumerate(['Rep', 'Years Active', 'Peak Annual Rev', 'Total Career Rev', 'Clients', 'Left', 'Revenue After = $0'], 1):
    ws5.cell(17, c, h)
style_header_row(ws5, 17, 7)

lost = [
    ('John Luedke', '24yr', 1018083, 12588833, 407, 2019),
    ('Gina Ellis', '20yr', 1308024, 10109144, 147, 2015),
    ('Ryan Gades', '17yr', 1179745, 5250471, 99, 2019),
    ('Maria Mercanti', '7yr', 1243944, 3945279, 207, 2024),
    ('Fausto Guillermo', '3yr', 1177815, 2521558, 99, 2024),
    ('Drew Brady', '18yr', 508955, 2388246, 66, 2018),
    ('Catherine Myers', '21yr', 456100, 2063037, 30, 2018),
    ('Courtney Blue', '3yr', 478997, 1250695, 20, 2010),
]
for i, (name, yrs, peak, total, clients, left) in enumerate(lost):
    row = 18 + i
    style_data_cell(ws5, row, 1).value = name
    style_data_cell(ws5, row, 2).value = yrs
    style_data_cell(ws5, row, 3, money_fmt).value = peak
    style_data_cell(ws5, row, 4, money_fmt).value = total
    style_data_cell(ws5, row, 5, '#,##0').value = clients
    style_data_cell(ws5, row, 6).value = left
    ws5.cell(row, 6).number_format = '0'
    style_data_cell(ws5, row, 7).value = '$0'
    ws5.cell(row, 7).fill = red_fill

for col in range(1, 8):
    ws5.column_dimensions[get_column_letter(col)].width = 18

# ============================================================
# SHEET 6: Inflation Reference
# ============================================================
ws6 = wb.create_sheet('Inflation Reference')
ws6.sheet_properties.tabColor = '666666'

ws6['A1'] = 'CPI-U Multipliers for Inflation Adjustment'
ws6['A1'].font = title_font
ws6.merge_cells('A1:D1')
ws6['A2'] = 'Source: BLS CPI-U Annual Averages. 2025-2026 estimated.'
ws6['A2'].font = Font(name='Arial', italic=True, size=9, color='666666')

for c, h in enumerate(['Year', 'CPI-U', 'Multiplier to 2026$', 'Example: $100K in that year = '], 1):
    ws6.cell(4, c, h)
style_header_row(ws6, 4, 4)

for i, yr in enumerate(sorted(cpi.keys())):
    row = 5 + i
    mult = cpi_2026 / cpi[yr]
    style_data_cell(ws6, row, 1).value = yr
    ws6.cell(row, 1).number_format = '0'
    style_data_cell(ws6, row, 2, '0.0').value = cpi[yr]
    style_data_cell(ws6, row, 3, '0.00x').value = mult
    style_data_cell(ws6, row, 4, money_fmt).value = round(100000 * mult)

for col in range(1, 5):
    ws6.column_dimensions[get_column_letter(col)].width = 22

# ============================================================
# Save
# ============================================================
output_path = r'C:\tmp\tsi-sales-rep-benchmarks.xlsx'
wb.save(output_path)
print(f'Saved to {output_path}')
print('Sheets:', [s.title for s in wb.worksheets])
