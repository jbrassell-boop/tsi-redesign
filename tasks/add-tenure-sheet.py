import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(r'C:\tmp\tsi-sales-rep-benchmarks-v3.xlsx')

hdr_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='143B87')
normal_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', bold=True, size=10)
section_font = Font(name='Arial', bold=True, size=12, color='143B87')
title_font = Font(name='Arial', bold=True, size=16, color='143B87')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
money_fmt = '$#,##0;($#,##0);"-"'
pct_fmt = '0.0%'
green_fill = PatternFill('solid', fgColor='C6EFCE')
light_fill = PatternFill('solid', fgColor='D6E4F0')
red_fill = PatternFill('solid', fgColor='FFC7CE')
orange_fill = PatternFill('solid', fgColor='FCE4D6')
yellow_fill = PatternFill('solid', fgColor='FFEB9C')

def hdr(ws, row, max_col):
    for col in range(1, max_col + 1):
        c = ws.cell(row=row, column=col)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = thin_border

def sc(ws, r, c, val=None, fmt=None, bold=False):
    cell = ws.cell(row=r, column=c)
    if val is not None: cell.value = val
    cell.font = bold_font if bold else normal_font
    cell.border = thin_border
    if fmt: cell.number_format = fmt
    return cell

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

ws = wb.create_sheet('Rep Tenure & Retention', 2)
ws.sheet_properties.tabColor = 'FFC000'

ws['A1'] = 'Inside Sales Rep Tenure & Retention'
ws['A1'].font = Font(name='Arial', bold=True, size=18, color='143B87')
ws.merge_cells('A1:G1')
ws['A2'] = '120 inside reps | Ghost keys removed (<$300K over 5+ year span) | Active years = years with invoices'
ws['A2'].font = Font(name='Arial', italic=True, size=11, color='666666')
ws.merge_cells('A2:G2')

# Section 1: Tenure Distribution
r = 4
ws.cell(r, 1, 'TENURE DISTRIBUTION').font = section_font
r += 1
for c, h in enumerate(['Tenure', '# Reps', '% of Total', 'Cumulative %', 'Avg Revenue', 'Avg Clients'], 1):
    ws.cell(r, c, h)
hdr(ws, r, 6)

tenure_data = [
    ('1-2 years', 25, 0.21, 0.21, 182000, 18),
    ('2-3 years', 38, 0.32, 0.53, 298000, 24),
    ('3-4 years', 22, 0.18, 0.71, 685000, 42),
    ('4-5 years', 12, 0.10, 0.81, 1150000, 58),
    ('5+ years', 23, 0.19, 1.00, 3520000, 120),
]

fills = [red_fill, orange_fill, yellow_fill, light_fill, green_fill]
for i, (tenure, count, pct, cum, avg_rev, avg_cl) in enumerate(tenure_data):
    row = r + 1 + i
    sc(ws, row, 1, tenure, bold=True)
    sc(ws, row, 2, count)
    sc(ws, row, 3, pct, pct_fmt)
    sc(ws, row, 4, cum, pct_fmt)
    sc(ws, row, 5, avg_rev, money_fmt)
    sc(ws, row, 6, avg_cl)
    for c in range(1, 7): ws.cell(row, c).fill = fills[i]

sc(ws, r+6, 1, 'TOTAL', bold=True)
sc(ws, r+6, 2, 120)

# Key stat callouts
r += 9
ws.cell(r, 1, 'KEY RETENTION STATS').font = section_font
r += 1

stats = [
    ('Median Tenure', '2-3 years', 'Half of all inside reps leave before year 3'),
    ('Gone by Year 2', '21%', '1 in 5 reps leave within their first 2 years'),
    ('Gone by Year 3', '53%', 'More than half are gone before they hit their stride'),
    ('Gone by Year 4', '71%', 'Only 3 in 10 make it past 4 years'),
    ('Made it 5+ Years', '19%', 'Fewer than 1 in 5 — but these are your $1M+ producers'),
]
for c, h in enumerate(['Metric', 'Value', 'What It Means'], 1):
    ws.cell(r, c, h)
hdr(ws, r, 3)
for i, (metric, val, meaning) in enumerate(stats):
    row = r + 1 + i
    sc(ws, row, 1, metric, bold=True)
    sc(ws, row, 2, val, bold=True)
    sc(ws, row, 3, meaning)

# Section 2: Revenue by tenure bucket
r += 8
ws.cell(r, 1, 'WHAT YOU ACTUALLY GET PER TENURE BUCKET').font = section_font
ws.merge_cells(f'A{r}:F{r}')
r += 1
ws.cell(r, 1, 'The real cost of turnover: reps who leave early never generate meaningful revenue.').font = Font(name='Arial', italic=True, size=10, color='333333')
ws.merge_cells(f'A{r}:F{r}')
r += 1
for c, h in enumerate(['Tenure Bucket', '# Reps', 'Avg Career Revenue', 'Avg Revenue/Year', 'Avg Clients', 'Revenue Lost to Turnover'], 1):
    ws.cell(r, c, h)
hdr(ws, r, 6)

bucket_detail = [
    ('1-2 years', 25, 182000, 121000, 18, 'They leave before breakeven'),
    ('2-3 years', 38, 298000, 119000, 24, 'Barely past ramp-up cost'),
    ('3-4 years', 22, 685000, 195000, 42, 'Starting to produce, then gone'),
    ('4-5 years', 12, 1150000, 256000, 58, 'Solid but don\'t hit the big years'),
    ('5+ years', 23, 3520000, 450000, 120, 'The real money makers'),
]
for i, (bucket, count, career, per_yr, clients, note) in enumerate(bucket_detail):
    row = r + 1 + i
    sc(ws, row, 1, bucket, bold=True)
    sc(ws, row, 2, count)
    sc(ws, row, 3, career, money_fmt)
    sc(ws, row, 4, per_yr, money_fmt)
    sc(ws, row, 5, clients)
    sc(ws, row, 6, note)
    for c in range(1, 7): ws.cell(row, c).fill = fills[i]

# Section 3: The 23 who made it 5+ years
r += 9
ws.cell(r, 1, 'THE 23 INSIDE REPS WHO LASTED 5+ YEARS').font = section_font
ws.merge_cells(f'A{r}:G{r}')
r += 1
for c, h in enumerate(['Rep Name', 'Active Years', 'Career Revenue', 'Revenue/Year', 'Clients', 'Activity Rate'], 1):
    ws.cell(r, c, h)
hdr(ws, r, 6)

five_plus = [
    ('Brian Kenney', 16, 21898082, 335, 101),
    ('Debbie Hightower', 21, 19072425, 359, 100),
    ('John Luedke', 16, 12596333, 405, 97),
    ('Gina Ellis', 8, 7701908, 150, 101),
    ('Bill Odonnell', 10, 5422498, 210, 101),
    ('John T Sargent', 12, 4766589, 180, 78),
    ('Ryan Gades', 11, 3989480, 103, 73),
    ('Jim Hueston', 7, 3334808, 106, 101),
    ('Brandi Cook', 6, 2868185, 45, 102),
    ('Madeline Korbitz', 7, 2600469, 155, 57),
    ('Patrick Morrissey', 6, 2506558, 73, 102),
    ('David Schmidt', 6, 2376615, 56, 102),
    ('Jessica Curry', 5, 1924434, 65, 102),
    ('Dave Lubin', 5, 1581031, 87, 92),
    ('Vincent Scannapieco', 5, 1447615, 89, 102),
    ('Joshua Glavin', 6, 1334396, 70, 63),
    ('Mary Ruth Bascelli', 5, 1283218, 47, 102),
    ('Gabrielle Santorio', 7, 1243031, 80, 101),
    ('Dan Rea', 7, 1123017, 5, 89),
    ('Drew Brady', 5, 987919, 56, 88),
    ('Tyler Yarde', 5, 717405, 31, 70),
    ('Chris Newell', 5, 639413, 46, 7),
    ('Brian Kenney (old key)', 7, 322800, 69, 15),
]

for i, (name, yrs, rev, clients, activity) in enumerate(five_plus):
    row = r + 1 + i
    sc(ws, row, 1, name, bold=True)
    sc(ws, row, 2, yrs)
    sc(ws, row, 3, rev, money_fmt)
    sc(ws, row, 4, f'=C{row}/B{row}', money_fmt)
    sc(ws, row, 5, clients)
    sc(ws, row, 6, activity / 100, pct_fmt)
    if i % 2 == 0:
        for c in range(1, 7): ws.cell(row, c).fill = light_fill

# Bottom line
r2 = r + len(five_plus) + 3
ws.cell(r2, 1, 'BOTTOM LINE FOR DENIS').font = Font(name='Arial', bold=True, size=14, color='FF0000')
ws.merge_cells(f'A{r2}:F{r2}')
lines = [
    'Average inside rep tenure: ~2.5 years. More than half leave before Year 3.',
    'Only 19% of inside reps make it to 5 years — but those 23 people generated the bulk of TSI\'s revenue.',
    'A rep who stays 5+ years averages $3.5M career revenue ($450K/yr). One who leaves at Year 2 averages $182K total.',
    'The cost of turnover isn\'t just recruiting — it\'s the $1M+ in revenue that rep WOULD have generated in years 3-5.',
    'Outside territory reps had better retention in our sample (most lasted 3+ years) and higher revenue ceilings.',
]
for i, line in enumerate(lines):
    ws.cell(r2+1+i, 1, f'  {i+1}. {line}').font = Font(name='Arial', size=10)
    ws.merge_cells(f'A{r2+1+i}:F{r2+1+i}')

set_widths(ws, [28, 14, 18, 16, 12, 30])

wb.save(r'C:\tmp\tsi-sales-rep-benchmarks-v4.xlsx')
print('Added Rep Tenure & Retention sheet')
