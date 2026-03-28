import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

hdr_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='143B87')
normal_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', bold=True, size=10)
section_font = Font(name='Arial', bold=True, size=12, color='143B87')
title_font = Font(name='Arial', bold=True, size=16, color='143B87')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
thick_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
money_fmt = '$#,##0;($#,##0);"-"'
pct_fmt = '0.0%'
num_fmt = '#,##0'
green_fill = PatternFill('solid', fgColor='C6EFCE')
light_fill = PatternFill('solid', fgColor='D6E4F0')
yellow_fill = PatternFill('solid', fgColor='FFFF00')
red_fill = PatternFill('solid', fgColor='FFC7CE')
orange_fill = PatternFill('solid', fgColor='FCE4D6')
total_fill = PatternFill('solid', fgColor='143B87')
total_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
gap_font = Font(name='Arial', bold=True, size=14, color='FF0000')

def hdr(ws, row, max_col):
    for col in range(1, max_col + 1):
        c = ws.cell(row=row, column=col)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', wrap_text=True); c.border = thin_border

def sc(ws, r, c, val=None, fmt=None, bold=False):
    cell = ws.cell(row=r, column=c)
    if val is not None: cell.value = val
    cell.font = bold_font if bold else normal_font; cell.border = thin_border
    if fmt: cell.number_format = fmt
    return cell

def total_row_style(ws, row, max_col):
    for c in range(1, max_col + 1):
        ws.cell(row, c).font = total_font; ws.cell(row, c).fill = total_fill; ws.cell(row, c).border = thin_border

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# SHEET 1: Revenue Goal Planner (the solver)
# ============================================================
ws = wb.active
ws.title = 'Revenue Goal Planner'
ws.sheet_properties.tabColor = 'FF0000'

ws['A1'] = 'TSI Revenue Goal Planner'
ws['A1'].font = Font(name='Arial', bold=True, size=20, color='143B87')
ws.merge_cells('A1:H1')
ws['A2'] = 'Set your goal. See what you need to hire. All yellow cells are editable.'
ws['A2'].font = Font(name='Arial', italic=True, size=11, color='666666')
ws.merge_cells('A2:H2')

# ─── INPUTS ───
r = 4
ws.cell(r, 1, 'REVENUE GOAL').font = Font(name='Arial', bold=True, size=14, color='FF0000')
c = sc(ws, r, 2, 20000000, money_fmt)
c.fill = yellow_fill; c.border = thick_border; c.font = Font(name='Arial', bold=True, size=16, color='FF0000')
goal_cell = f'B{r}'
goal_row = r

r += 2
ws.cell(r, 1, 'NEW HIRE ASSUMPTIONS').font = section_font
ws.merge_cells(f'A{r}:H{r}')
r += 1
ws.cell(r, 1, 'Based on 25 years of TSI invoice data (200K+ invoices, 120 reps). Edit yellow to adjust.').font = Font(name='Arial', italic=True, size=9, color='666666')
ws.merge_cells(f'A{r}:H{r}')
r += 1

# Assumption labels + editable values
for c_idx, (label, val, fmt) in enumerate([
    ('Outside Rep Yr1 Revenue', 348374, money_fmt),
    ('Inside Rep Yr1 Revenue', 202000, money_fmt),
    ('YoY Growth Rate', 0.30, pct_fmt),
    ('Outside Rep Comp/Yr', 100000, money_fmt),
    ('Inside Rep Comp/Yr', 65000, money_fmt),
], 1):
    arow = r + (c_idx - 1)
    sc(ws, arow, 1, label, bold=True)
    c = sc(ws, arow, 2, val, fmt)
    c.fill = yellow_fill; c.border = thick_border

outside_yr1_cell = f'B{r}'      # B8
inside_yr1_cell = f'B{r+1}'     # B9
growth_cell = f'B{r+2}'         # B10
outside_comp_cell = f'B{r+3}'   # B11
inside_comp_cell = f'B{r+4}'    # B12
assume_start = r
r += 6

# ─── CURRENT ROSTER ───
ws.cell(r, 1, 'CURRENT ROSTER').font = section_font
ws.merge_cells(f'A{r}:H{r}')
r += 1
ws.cell(r, 1, 'Revenue = average of 2024 + 2025 to smooth contract spikes. Edit to override.').font = Font(name='Arial', italic=True, size=9, color='666666')
ws.merge_cells(f'A{r}:H{r}')
r += 1

headers = ['Rep Name', 'Type', 'FFS Revenue', 'Contract Revenue', 'Total Revenue', 'Clients', 'Notes']
for c, h in enumerate(headers, 1): ws.cell(r, c, h)
hdr(ws, r, 7)
roster_hdr = r
r += 1

# Current reps: (name, type, avg_ffs_24_25, avg_con_24_25, clients, notes)
# Using average of 2024+2025 to smooth
current = [
    ('Brian Kenney', 'Inside', 3289768, 1844788, 76, 'Surgical Solutions contract'),
    ('Debbie Hightower', 'Inside', 1529992, 97092, 130, ''),
    ('Bernie DeLacy', 'Outside', 1172447, 577758, 46, 'Bayhealth contract (2024)'),
    ('Brandi Cook', 'Inside', 1058985, 756841, 28, '66% contract revenue'),
    ('Seamus Glavin', 'Inside', 870515, 269037, 61, ''),
    ('Jim Rygiel', 'Outside', 380467, 62142, 11, 'Ramping fast'),
    ('Timothy Reilly', 'Inside', 407527, 80283, 47, ''),
    ('Donald Haynes', 'Inside', 125999, 0, 47, 'Yr1 ramp'),
    ('Ryan George', 'Inside', 153895, 84374, 13, ''),
]

bsc = [
    ('BSC Sole Flex', 'Program', 0, 687396, 0, ''),
    ('BSC CAP', 'Program', 0, 672138, 0, ''),
    ('BSC CAP-S', 'Program', 0, 422831, 0, ''),
    ('BSC CAP-M', 'Program', 0, 305155, 0, ''),
    ('EC Legacy', 'Program', 0, 111131, 0, ''),
]

roster_first = r
for name, rtype, ffs, con, clients, notes in current:
    sc(ws, r, 1, name, bold=True)
    sc(ws, r, 2, rtype)
    c = sc(ws, r, 3, ffs, money_fmt); c.fill = yellow_fill
    c = sc(ws, r, 4, con, money_fmt); c.fill = yellow_fill
    sc(ws, r, 5, f'=C{r}+D{r}', money_fmt)
    sc(ws, r, 6, clients, num_fmt)
    sc(ws, r, 7, notes)
    fill = green_fill if rtype == 'Outside' else light_fill
    ws.cell(r, 1).fill = fill; ws.cell(r, 2).fill = fill
    r += 1

for name, rtype, ffs, con, clients, notes in bsc:
    sc(ws, r, 1, name, bold=True)
    sc(ws, r, 2, rtype)
    c = sc(ws, r, 3, ffs, money_fmt); c.fill = yellow_fill
    c = sc(ws, r, 4, con, money_fmt); c.fill = yellow_fill
    sc(ws, r, 5, f'=C{r}+D{r}', money_fmt)
    sc(ws, r, 6, clients, num_fmt)
    sc(ws, r, 7, notes)
    for c_idx in range(1, 3): ws.cell(r, c_idx).fill = orange_fill
    r += 1

roster_last = r - 1

# Current total
sc(ws, r, 1, 'CURRENT ROSTER TOTAL', bold=True)
for col in [3, 4, 5]:
    cl = get_column_letter(col)
    sc(ws, r, col, f'=SUM({cl}{roster_first}:{cl}{roster_last})', money_fmt)
total_row_style(ws, r, 7)
current_total_row = r
r += 1

# Gap before new hires
sc(ws, r, 1, 'GAP TO GOAL (before new hires)', bold=True)
sc(ws, r, 5, f'={goal_cell}-E{current_total_row}', money_fmt).font = gap_font
ws.cell(r, 1).font = gap_font
gap_row = r
r += 2

# ─── WHAT YOU NEED TO HIRE ───
ws.cell(r, 1, 'WHAT YOU NEED TO HIRE').font = Font(name='Arial', bold=True, size=14, color='143B87')
ws.merge_cells(f'A{r}:H{r}')
r += 1
ws.cell(r, 1, 'Auto-calculated: enough reps to close the gap. Change mix in yellow cells.').font = Font(name='Arial', italic=True, size=9, color='666666')
ws.merge_cells(f'A{r}:H{r}')
r += 1

# Hire mix inputs
sc(ws, r, 1, '# Outside Reps to Hire:', bold=True)
c = sc(ws, r, 2, 2, num_fmt); c.fill = yellow_fill; c.border = thick_border
outside_hire_cell = f'B{r}'
r += 1
sc(ws, r, 1, '# Inside Reps to Hire:', bold=True)
c = sc(ws, r, 2, 3, num_fmt); c.fill = yellow_fill; c.border = thick_border
inside_hire_cell = f'B{r}'
r += 2

# New hire projection table
for c, h in enumerate(['', 'Type', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5', '5-Yr Total'], 1):
    ws.cell(r, c, h)
hdr(ws, r, 8)
hire_hdr = r
r += 1

# Outside hire rows
for i in range(5):
    sc(ws, r, 1, f'Outside Rep {i+1}', bold=True)
    sc(ws, r, 2, 'Outside')
    # Show revenue only if this rep number <= hire count
    for yr in range(1, 6):
        col = yr + 2
        if yr == 1:
            sc(ws, r, col, f'=IF({i+1}<={outside_hire_cell},{outside_yr1_cell},0)', money_fmt)
        else:
            prev = get_column_letter(col - 1)
            sc(ws, r, col, f'=IF({i+1}<={outside_hire_cell},{prev}{r}*(1+{growth_cell}),0)', money_fmt)
    sc(ws, r, 8, f'=SUM(C{r}:G{r})', money_fmt)
    for c_idx in range(1, 9): ws.cell(r, c_idx).fill = green_fill if (i % 2 == 0) else PatternFill('solid', fgColor='E2EFDA')
    r += 1

outside_first = hire_hdr + 1
outside_last = r - 1
r += 1  # gap

# Inside hire rows
for i in range(5):
    sc(ws, r, 1, f'Inside Rep {i+1}', bold=True)
    sc(ws, r, 2, 'Inside')
    for yr in range(1, 6):
        col = yr + 2
        if yr == 1:
            sc(ws, r, col, f'=IF({i+1}<={inside_hire_cell},{inside_yr1_cell},0)', money_fmt)
        else:
            prev = get_column_letter(col - 1)
            sc(ws, r, col, f'=IF({i+1}<={inside_hire_cell},{prev}{r}*(1+{growth_cell}),0)', money_fmt)
    sc(ws, r, 8, f'=SUM(C{r}:G{r})', money_fmt)
    for c_idx in range(1, 9): ws.cell(r, c_idx).fill = light_fill if (i % 2 == 0) else PatternFill('solid', fgColor='E8F0FE')
    r += 1

inside_first = outside_last + 2
inside_last = r - 1

# New hire totals
sc(ws, r, 1, 'NEW HIRES TOTAL', bold=True)
for col in range(3, 9):
    cl = get_column_letter(col)
    sc(ws, r, col, f'=SUM({cl}{outside_first}:{cl}{outside_last})+SUM({cl}{inside_first}:{cl}{inside_last})', money_fmt)
total_row_style(ws, r, 8)
new_total_row = r
r += 2

# ─── FINAL PROJECTION ───
ws.cell(r, 1, 'PROJECTION').font = Font(name='Arial', bold=True, size=14, color='143B87')
ws.merge_cells(f'A{r}:H{r}')
r += 1

for c, h in enumerate(['', '', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'], 1):
    ws.cell(r, c, h)
hdr(ws, r, 7)
r += 1

# Current (assumed flat)
sc(ws, r, 1, 'Current Roster (flat)', bold=True)
for col in range(3, 8):
    sc(ws, r, col, f'=E{current_total_row}', money_fmt)
for c_idx in range(1, 8): ws.cell(r, c_idx).fill = light_fill
r += 1

# New hires
sc(ws, r, 1, 'New Hire Revenue', bold=True)
for col in range(3, 8):
    cl = get_column_letter(col)
    sc(ws, r, col, f'={cl}{new_total_row}', money_fmt)
for c_idx in range(1, 8): ws.cell(r, c_idx).fill = green_fill
r += 1

# Combined
sc(ws, r, 1, 'PROJECTED TOTAL', bold=True)
for col in range(3, 8):
    cl = get_column_letter(col)
    sc(ws, r, col, f'={cl}{r-2}+{cl}{r-1}', money_fmt)
total_row_style(ws, r, 7)
proj_row = r
r += 1

# Goal
sc(ws, r, 1, 'Revenue Goal', bold=True)
for col in range(3, 8):
    sc(ws, r, col, f'={goal_cell}', money_fmt)
ws.cell(r, 1).fill = yellow_fill
r += 1

# Gap
sc(ws, r, 1, 'GAP (negative = short)', bold=True)
ws.cell(r, 1).font = gap_font
for col in range(3, 8):
    cl = get_column_letter(col)
    cell = sc(ws, r, col, f'={cl}{proj_row}-{cl}{r-1}', money_fmt)
    cell.font = gap_font
r += 2

# ─── INVESTMENT / ROI ───
ws.cell(r, 1, 'INVESTMENT').font = section_font
r += 1
for c, h in enumerate(['', '', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'], 1):
    ws.cell(r, c, h)
hdr(ws, r, 7)
r += 1

sc(ws, r, 1, 'New Hire Comp Cost', bold=True)
for col in range(3, 8):
    sc(ws, r, col, f'={outside_hire_cell}*{outside_comp_cell}+{inside_hire_cell}*{inside_comp_cell}', money_fmt)
comp_row = r
r += 1

sc(ws, r, 1, 'New Hire Revenue', bold=True)
for col in range(3, 8):
    cl = get_column_letter(col)
    sc(ws, r, col, f'={cl}{new_total_row}', money_fmt)
r += 1

sc(ws, r, 1, 'Net Return', bold=True)
for col in range(3, 8):
    cl = get_column_letter(col)
    cell = sc(ws, r, col, f'={cl}{r-1}-{cl}{comp_row}', money_fmt)
    cell.font = Font(name='Arial', bold=True, size=11)
r += 1

sc(ws, r, 1, 'ROI', bold=True)
for col in range(3, 8):
    cl = get_column_letter(col)
    sc(ws, r, col, f'=IF({cl}{comp_row}>0,{cl}{r-1}/{cl}{comp_row},0)', pct_fmt)

set_widths(ws, [28, 12, 16, 16, 16, 16, 16, 16])

# ============================================================
# SHEET 2: Inside vs Outside History (same as v5)
# ============================================================
ws2 = wb.create_sheet('Inside vs Outside')
ws2.sheet_properties.tabColor = '143B87'
ws2['A1'] = 'Historical: Inside vs Outside Territory Rep Performance'
ws2['A1'].font = title_font
ws2.merge_cells('A1:G1')
ws2['A2'] = '8 outside territory reps | 100+ inside reps | 25 years of invoice data | FFS + Contract'
ws2['A2'].font = Font(name='Arial', italic=True, size=10, color='666666')

r = 4
ws2.cell(r, 1, 'TOTAL REVENUE BY TENURE YEAR').font = section_font
r += 1
for c, h in enumerate(['', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'], 1): ws2.cell(r, c, h)
hdr(ws2, r, 6)

data = [
    ('Outside Territory', [348374, 404261, 491083, 814695, 1455687], green_fill, '8 reps'),
    ('Inside / Phone', [136078, 218961, 301268, 312861, 390956], light_fill, '101+ reps'),
]
for i, (label, vals, fill, note) in enumerate(data):
    row = r + 1 + i
    sc(ws2, row, 1, f'{label} ({note})', bold=True)
    for c, v in enumerate(vals, 2): sc(ws2, row, c, v, money_fmt)
    for c in range(1, 7): ws2.cell(row, c).fill = fill

r += 4
ws2.cell(r, 1, 'FFS vs CONTRACT BREAKDOWN').font = section_font
r += 1
for c, h in enumerate(['', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'], 1): ws2.cell(r, c, h)
hdr(ws2, r, 6)

bd = [
    ('Outside FFS', [348374, 404261, 491083, 814695, 1455687], green_fill),
    ('Outside Contract', [0, 0, 0, 0, 0], green_fill),
    ('Inside FFS', [127025, 174756, 223942, 239933, 302433], light_fill),
    ('Inside Contract', [9053, 44204, 77325, 72928, 88523], light_fill),
]
for i, (label, vals, fill) in enumerate(bd):
    row = r + 1 + i
    sc(ws2, row, 1, label, bold=True)
    for c, v in enumerate(vals, 2): sc(ws2, row, c, v, money_fmt)
    for c in range(1, 7): ws2.cell(row, c).fill = fill

r += 7
ws2.cell(r, 1, 'NET NEW REVENUE').font = section_font
r += 1
for c, h in enumerate(['', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'], 1): ws2.cell(r, c, h)
hdr(ws2, r, 6)

nn = [
    ('Outside Net New $', [24178, 21232, 204770, 524414, 1083019], green_fill, money_fmt),
    ('Outside Net New %', [0.07, 0.05, 0.42, 0.64, 0.74], green_fill, pct_fmt),
    ('Inside Net New $', [28212, 57618, 95843, 102939, 129510], light_fill, money_fmt),
    ('Inside Net New %', [0.21, 0.26, 0.32, 0.33, 0.33], light_fill, pct_fmt),
]
for i, (label, vals, fill, fmt) in enumerate(nn):
    row = r + 1 + i
    sc(ws2, row, 1, label, bold=True)
    for c, v in enumerate(vals, 2): sc(ws2, row, c, v, fmt)
    for c in range(1, 7): ws2.cell(row, c).fill = fill

r += 7
ws2.cell(r, 1, 'UNIQUE CLIENTS').font = section_font
r += 1
for c, h in enumerate(['', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'], 1): ws2.cell(r, c, h)
hdr(ws2, r, 6)
sc(ws2, r+1, 1, 'Outside Clients', bold=True)
for c, v in enumerate([23, 17, 15, 8, 14], 2): sc(ws2, r+1, c, v)
for c in range(1, 7): ws2.cell(r+1, c).fill = green_fill
sc(ws2, r+2, 1, 'Inside Clients', bold=True)
for c, v in enumerate([16, 19, 23, 22, 25], 2): sc(ws2, r+2, c, v)
for c in range(1, 7): ws2.cell(r+2, c).fill = light_fill

r += 5
ws2.cell(r, 1, 'RETENTION (Inside Reps)').font = section_font
r += 1
for c, h in enumerate(['Tenure', '# Reps', '% Gone', 'Meaning'], 1): ws2.cell(r, c, h)
hdr(ws2, r, 4)
ret = [
    ('< 3 years', 63, '53%', 'More than half leave before hitting stride'),
    ('3-5 years', 34, '28%', 'Producing but miss peak years'),
    ('5+ years', 23, '19%', 'The $1M+ producers — only 1 in 5 make it'),
]
for i, (t, n, pct, note) in enumerate(ret):
    row = r + 1 + i
    sc(ws2, row, 1, t, bold=True); sc(ws2, row, 2, n); sc(ws2, row, 3, pct); sc(ws2, row, 4, note)

r += 6
ws2.cell(r, 1, 'KEY TAKEAWAYS FOR DENIS').font = Font(name='Arial', bold=True, size=14, color='FF0000')
insights = [
    'Outside reps: higher ceiling ($800K+ by yr4), but 100% FFS, slower net new, bigger inherited book needed.',
    'Inside reps: lower ceiling ($390K yr5), but add contracts ($45-89K/yr by yr2-5) and source net new faster.',
    'Contract revenue = 36% of company total ($6.6M of $18.4M in 2025). Contracts are an inside rep play.',
    '53% of inside reps leave within 3 years. Only 19% last 5+. Budget for turnover.',
    'The question isn\'t inside OR outside — it\'s the right MIX to hit the revenue goal.',
]
for i, text in enumerate(insights):
    ws2.cell(r+1+i, 1, f'  {i+1}. {text}').font = Font(name='Arial', size=10)
    ws2.merge_cells(f'A{r+1+i}:F{r+1+i}')

set_widths(ws2, [34, 14, 14, 14, 14, 14])

# ============================================================
# SHEET 3: Outside Rep Profiles
# ============================================================
ws3 = wb.create_sheet('Outside Reps')
ws3.sheet_properties.tabColor = '70AD47'
ws3['A1'] = 'Outside Territory Rep Profiles (8 confirmed)'
ws3['A1'].font = title_font
for c, h in enumerate(['Rep', 'Territory', 'Active', 'Career FFS', 'Career Contract', 'Total', 'Clients', 'Mix'], 1): ws3.cell(3, c, h)
hdr(ws3, 3, 8)
reps = [
    ('Lindsey Davis', 'MD/DC', '2005-2010', 6548594, 0, 19, '54% Flex, 28% Rigid'),
    ('Eric Schwarzel', 'PA/NJ/DE', '2007-2010', 1263399, 0, 38, '35% Flex, 19% Rigid'),
    ('Courtney Blue', 'OR/TX', '2007-2010', 1250695, 0, 20, '47% Flex, 37% Instr'),
    ('Danielle Penge', 'OR/NJ', '2007-2010', 877655, 0, 27, '65% Flex, 17% Rigid'),
    ('Michael Woessner', 'NJ/OR', '2005-2009', 885795, 0, 23, '78% Flex, 11% Rigid'),
    ('Kara Klund', 'IL/TX', '2006-2008', 906112, 0, 20, '33% F, 25% R, 34% Cam'),
    ('Melissa Fox', 'MD/DC/VA', '2016-2018', 407129, 319586, 9, '72% Flex, 19% Rigid'),
    ('Bernie DeLacy', 'FL/PA', '~2023+', 371597, 1870094, 46, 'Mixed'),
]
for i, (name, terr, active, ffs, con, cl, mix) in enumerate(reps):
    row = 4 + i
    sc(ws3, row, 1, name, bold=True); sc(ws3, row, 2, terr); sc(ws3, row, 3, active)
    sc(ws3, row, 4, ffs, money_fmt); sc(ws3, row, 5, con, money_fmt)
    sc(ws3, row, 6, f'=D{row}+E{row}', money_fmt); sc(ws3, row, 7, cl); sc(ws3, row, 8, mix)
    if i % 2 == 0:
        for c in range(1, 9): ws3.cell(row, c).fill = light_fill
set_widths(ws3, [20, 14, 14, 16, 16, 16, 10, 30])

# ============================================================
# SHEET 4: Company Revenue
# ============================================================
ws4 = wb.create_sheet('Company Revenue')
ws4.sheet_properties.tabColor = '4472C4'
ws4['A1'] = 'TSI Revenue Summary'
ws4['A1'].font = title_font
for c, h in enumerate(['Year', 'FFS Revenue', 'Contract Revenue', 'Total', 'Contract %'], 1): ws4.cell(3, c, h)
hdr(ws4, 3, 5)
for i, (yr, ffs, con) in enumerate([(2024, 8958775, 5323596), (2025, 11771287, 6613925)]):
    row = 4 + i
    sc(ws4, row, 1, yr); ws4.cell(row, 1).number_format = '0'
    sc(ws4, row, 2, ffs, money_fmt); sc(ws4, row, 3, con, money_fmt)
    sc(ws4, row, 4, f'=B{row}+C{row}', money_fmt); sc(ws4, row, 5, f'=C{row}/D{row}', pct_fmt)
sc(ws4, 6, 1, 'Average', bold=True)
for col in range(2, 5): sc(ws4, 6, col, f'=AVERAGE({get_column_letter(col)}4:{get_column_letter(col)}5)', money_fmt)
sc(ws4, 6, 5, f'=C6/D6', pct_fmt)
ws4.cell(8, 1, '2024-2025 average used for current roster baseline to smooth contract spikes.').font = Font(name='Arial', italic=True, size=9, color='666666')
set_widths(ws4, [10, 16, 16, 16, 12])

# Save
out = r'C:\tmp\tsi-sales-rep-planner-v6.xlsx'
wb.save(out)
print(f'Saved to {out}')
print('Sheets:', [s.title for s in wb.worksheets])
