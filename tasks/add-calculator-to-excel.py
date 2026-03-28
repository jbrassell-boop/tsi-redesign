import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(r'C:\tmp\tsi-sales-rep-benchmarks.xlsx')

hdr_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='143B87')
normal_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', bold=True, size=10)
section_font = Font(name='Arial', bold=True, size=12, color='143B87')
title_font = Font(name='Arial', bold=True, size=16, color='143B87')
input_fill = PatternFill('solid', fgColor='FFFF00')
total_fill = PatternFill('solid', fgColor='143B87')
total_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
light_blue = PatternFill('solid', fgColor='D6E4F0')
green_fill = PatternFill('solid', fgColor='C6EFCE')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
thick_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
money_fmt = '$#,##0;($#,##0);"-"'

def sc(ws, r, c, val=None, fmt=None, bold=False):
    cell = ws.cell(row=r, column=c)
    if val is not None: cell.value = val
    cell.font = bold_font if bold else normal_font
    cell.border = thin_border
    if fmt: cell.number_format = fmt
    return cell

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

ws = wb.create_sheet('Revenue Planner', 1)
ws.sheet_properties.tabColor = 'FF0000'

ws['A1'] = 'Sales Revenue Headcount Planner'
ws['A1'].font = Font(name='Arial', bold=True, size=18, color='143B87')
ws.merge_cells('A1:H1')
ws['A2'] = 'Enter your revenue target and number of reps. See year-by-year buildup per rep and total.'
ws['A2'].font = Font(name='Arial', italic=True, size=11, color='666666')
ws.merge_cells('A2:H2')

# INPUTS section
ws.cell(4, 1, 'INPUTS').font = section_font

ws.cell(5, 1, 'Revenue Target:').font = bold_font
c = ws.cell(5, 2)
c.value = 3000000
c.number_format = money_fmt
c.fill = input_fill
c.border = thick_border
c.font = Font(name='Arial', bold=True, size=12)

ws.cell(6, 1, 'Number of New Reps:').font = bold_font
c = ws.cell(6, 2)
c.value = 5
c.number_format = '#,##0'
c.fill = input_fill
c.border = thick_border
c.font = Font(name='Arial', bold=True, size=12)

ws.cell(7, 1, 'Scenario:').font = bold_font
ws.cell(7, 2, 'Cold Territory (no inherited book)').font = Font(name='Arial', italic=True, size=10, color='FF0000')

ws.cell(8, 1, 'Rep Annual Comp (est):').font = bold_font
c = ws.cell(8, 2)
c.value = 80000
c.number_format = money_fmt
c.fill = input_fill
c.border = thick_border

# Assumptions
ws.cell(10, 1, 'ASSUMPTIONS (based on 25yr avg — edit yellow cells to adjust)').font = section_font
ws.merge_cells('A10:H10')

for c, h in enumerate(['', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'], 1):
    ws.cell(11, c, h)
style_header(ws, 11, 6)

# Net new revenue per rep (cold territory) - editable
ws.cell(12, 1, 'Revenue per Rep').font = bold_font
cold_vals = [44085, 74028, 146008, 176736, 257577]
for c, v in enumerate(cold_vals, 2):
    cell = sc(ws, 12, c, v, money_fmt)
    cell.fill = input_fill
    cell.border = thick_border

ws.cell(13, 1, 'Net New Clients per Rep').font = bold_font
client_vals = [9, 10, 12, 14, 18]
for c, v in enumerate(client_vals, 2):
    cell = sc(ws, 13, c, v, '#,##0')
    cell.fill = input_fill

# REP-BY-REP BUILDUP
ws.cell(15, 1, 'REP-BY-REP REVENUE BUILDUP').font = section_font
ws.merge_cells('A15:H15')
ws.cell(16, 1, 'Each row = one new rep. Revenue shown is cumulative as they ramp.').font = Font(name='Arial', italic=True, size=10, color='666666')
ws.merge_cells('A16:H16')

# Headers: Rep | Year 1 | Year 2 | Year 3 | Year 4 | Year 5 | 5-Yr Total
for c, h in enumerate(['Rep', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5', '5-Year Total', 'Clients (Yr5)'], 1):
    ws.cell(18, c, h)
style_header(ws, 18, 8)

# Generate 15 rep rows (formula-driven from the inputs)
max_reps = 15
for i in range(max_reps):
    row = 19 + i
    rep_num = i + 1
    # Rep label: "Rep 1", "Rep 2", etc. Grey out if > num reps input
    sc(ws, row, 1, f'Rep {rep_num}', bold=True)
    # Revenue per year — reference the assumptions row, but only show if rep_num <= B6
    for yr_col in range(2, 7):  # B through F = Year 1-5
        assumption_cell = f'{get_column_letter(yr_col)}$12'
        formula = f'=IF({rep_num}<=$B$6,{assumption_cell},0)'
        sc(ws, row, yr_col, formula, money_fmt)
    # 5-year total
    sc(ws, row, 7, f'=SUM(B{row}:F{row})', money_fmt)
    # Clients at year 5
    sc(ws, row, 8, f'=IF({rep_num}<=$B$6,F$13,0)', '#,##0')
    # Alternating row color
    if i % 2 == 0:
        for c in range(1, 9):
            ws.cell(row, c).fill = light_blue

# TOTALS ROW
total_row = 19 + max_reps
ws.cell(total_row, 1, 'TOTAL').font = total_font
ws.cell(total_row, 1).fill = total_fill
ws.cell(total_row, 1).border = thin_border
for col in range(2, 9):
    cell = ws.cell(total_row, col)
    col_letter = get_column_letter(col)
    cell.value = f'=SUM({col_letter}19:{col_letter}{total_row-1})'
    cell.number_format = money_fmt if col < 8 else '#,##0'
    cell.font = total_font
    cell.fill = total_fill
    cell.border = thin_border

# Target comparison
gap_row = total_row + 2
ws.cell(gap_row, 1, 'Revenue Target').font = bold_font
sc(ws, gap_row, 2, '=$B$5', money_fmt)
ws.cell(gap_row+1, 1, 'Total (5 years)').font = bold_font
sc(ws, gap_row+1, 2, f'=G{total_row}', money_fmt)
ws.cell(gap_row+2, 1, 'Gap to Target').font = Font(name='Arial', bold=True, size=11, color='FF0000')
sc(ws, gap_row+2, 2, f'=B{gap_row}-B{gap_row+1}', money_fmt)

# Investment analysis
inv_row = gap_row + 4
ws.cell(inv_row, 1, 'INVESTMENT ANALYSIS').font = section_font
ws.merge_cells(f'A{inv_row}:D{inv_row}')

ws.cell(inv_row+1, 1, '').font = bold_font
for c, h in enumerate(['', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'], 1):
    ws.cell(inv_row+1, c, h)
style_header(ws, inv_row+1, 6)

# Total comp cost
ws.cell(inv_row+2, 1, 'Total Comp Cost').font = bold_font
for yr_col in range(2, 7):
    sc(ws, inv_row+2, yr_col, f'=$B$6*$B$8', money_fmt)

# Total revenue
ws.cell(inv_row+3, 1, 'Total Revenue').font = bold_font
for yr_col in range(2, 7):
    col_letter = get_column_letter(yr_col)
    sc(ws, inv_row+3, yr_col, f'={col_letter}{total_row}', money_fmt)

# Net (Revenue - Cost)
ws.cell(inv_row+4, 1, 'Net (Rev - Comp)').font = Font(name='Arial', bold=True, size=11)
for yr_col in range(2, 7):
    col_letter = get_column_letter(yr_col)
    cell = sc(ws, inv_row+4, yr_col, f'={col_letter}{inv_row+3}-{col_letter}{inv_row+2}', money_fmt)
    cell.font = Font(name='Arial', bold=True, size=11)

# ROI
ws.cell(inv_row+5, 1, 'ROI').font = bold_font
for yr_col in range(2, 7):
    col_letter = get_column_letter(yr_col)
    sc(ws, inv_row+5, yr_col, f'={col_letter}{inv_row+4}/{col_letter}{inv_row+2}', '0.0%')

# Column widths
ws.column_dimensions['A'].width = 28
for c in range(2, 9):
    ws.column_dimensions[get_column_letter(c)].width = 16

# Print setup
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)

wb.save(r'C:\tmp\tsi-sales-rep-benchmarks.xlsx')
print('Added Revenue Planner sheet with calculator')
print('Sheets:', [s.title for s in wb.worksheets])
