import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

hdr_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='143B87')
normal = Font(name='Arial', size=10)
bold = Font(name='Arial', bold=True, size=10)
section = Font(name='Arial', bold=True, size=12, color='143B87')
thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
thick = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
money = '$#,##0;($#,##0);"-"'
pct = '0.0%'
yellow = PatternFill('solid', fgColor='FFFF00')
green = PatternFill('solid', fgColor='C6EFCE')
light = PatternFill('solid', fgColor='D6E4F0')
orange = PatternFill('solid', fgColor='FCE4D6')
navy = PatternFill('solid', fgColor='143B87')
navy_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
red_font = Font(name='Arial', bold=True, size=16, color='FF0000')
alt_fill = PatternFill('solid', fgColor='F2F2F2')

def hdr(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True); cell.border = thin

def sc(ws, r, c, val=None, fmt=None, b=False):
    cell = ws.cell(r, c)
    if val is not None: cell.value = val
    cell.font = bold if b else normal; cell.border = thin
    if fmt: cell.number_format = fmt
    return cell

def navy_row(ws, r, cols):
    for c in range(1, cols + 1):
        ws.cell(r, c).font = navy_font; ws.cell(r, c).fill = navy; ws.cell(r, c).border = thin

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# SHEET 1: Revenue Planner
# ============================================================
ws = wb.active
ws.title = 'Revenue Planner'
ws.sheet_properties.tabColor = '143B87'

ws['A1'] = 'TSI Revenue Planner'
ws['A1'].font = Font(name='Arial', bold=True, size=20, color='143B87')
ws.merge_cells('A1:G1')

# Goal
sc(ws, 3, 1, 'Revenue Goal', b=True).font = Font(name='Arial', bold=True, size=14)
c = sc(ws, 3, 2, 20000000, money)
c.fill = yellow; c.border = thick; c.font = Font(name='Arial', bold=True, size=18, color='FF0000')

# Current Roster
r = 5
ws.cell(r, 1, 'WHAT WE HAVE NOW').font = section
r += 1
for c, h in enumerate(['Rep', 'Type', 'Annual Revenue'], 1): ws.cell(r, c, h)
hdr(ws, r, 3)
r += 1

current = [
    ('Brian Kenney', 'Inside', 1455000),
    ('Debbie Hightower', 'Inside', 1627084),
    ('Bernie DeLacy', 'Outside', 1750205),
    ('Brandi Cook', 'Inside', 1815826),
    ('Seamus Glavin', 'Inside', 1139552),
    ('Jim Rygiel', 'Outside', 442609),
    ('Timothy Reilly', 'Inside', 487810),
    ('Donald Haynes', 'Inside', 125999),
    ('Ryan George', 'Inside', 238269),
    ('BSC Programs', 'Program', 2198651),
]

roster_first = r
for name, rtype, rev in current:
    sc(ws, r, 1, name, b=True)
    sc(ws, r, 2, rtype)
    c = sc(ws, r, 3, rev, money); c.fill = yellow
    fill = green if rtype == 'Outside' else orange if rtype == 'Program' else light
    ws.cell(r, 1).fill = fill; ws.cell(r, 2).fill = fill
    r += 1
roster_last = r - 1

sc(ws, r, 1, 'CURRENT TOTAL', b=True)
sc(ws, r, 3, f'=SUM(C{roster_first}:C{roster_last})', money)
navy_row(ws, r, 3)
current_total = r
r += 1

sc(ws, r, 1, 'GAP TO GOAL', b=True)
sc(ws, r, 3, f'=B3-C{current_total}', money).font = red_font
ws.cell(r, 1).font = Font(name='Arial', bold=True, size=14, color='FF0000')
r += 2

# Add Reps
ws.cell(r, 1, 'ADD REPS TO CLOSE THE GAP').font = section
r += 1
ws.cell(r, 1, 'Type "Outside" or "Inside". Revenue auto-fills. Clear Type to remove a rep.').font = Font(name='Arial', italic=True, size=9, color='666666')
ws.merge_cells(f'A{r}:G{r}')
r += 1

for c, h in enumerate(['New Rep', 'Type', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'], 1):
    ws.cell(r, c, h)
hdr(ws, r, 7)
hire_hdr = r
r += 1

# Benchmarks in hidden reference (row 55-56)
outside_bench = [356904, 400282, 209953, 814695, 1455687]
inside_bench = [145928, 146988, 222298, 273829, 364239]
for i, v in enumerate(outside_bench): sc(ws, 55, i + 1, v, money)
for i, v in enumerate(inside_bench): sc(ws, 56, i + 1, v, money)

hire_first = r
for i in range(10):
    sc(ws, r, 1, f'New Rep {i+1}')
    c = sc(ws, r, 2, 'Outside' if i < 2 else 'Inside' if i < 5 else '')
    c.fill = yellow
    for yr in range(1, 6):
        col = yr + 2
        ref = get_column_letter(yr)
        sc(ws, r, col, f'=IF(B{r}="Outside",{ref}55,IF(B{r}="Inside",{ref}56,0))', money)
    if i % 2 == 0:
        for c_idx in range(1, 8):
            if not ws.cell(r, c_idx).fill or ws.cell(r, c_idx).fill.fgColor is None:
                ws.cell(r, c_idx).fill = alt_fill
    r += 1
hire_last = r - 1

sc(ws, r, 1, 'NEW HIRES TOTAL', b=True)
for col in range(3, 8):
    cl = get_column_letter(col)
    sc(ws, r, col, f'=SUM({cl}{hire_first}:{cl}{hire_last})', money)
navy_row(ws, r, 7)
hire_total = r
r += 2

# The Answer
ws.cell(r, 1, 'THE ANSWER').font = Font(name='Arial', bold=True, size=16, color='143B87')
r += 1
for c, h in enumerate(['', '', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'], 1): ws.cell(r, c, h)
hdr(ws, r, 7)
r += 1

sc(ws, r, 1, 'Current Roster', b=True)
for col in range(3, 8): sc(ws, r, col, f'=C{current_total}', money)
for c in range(1, 8): ws.cell(r, c).fill = light
r += 1

sc(ws, r, 1, 'New Hire Revenue', b=True)
for col in range(3, 8): sc(ws, r, col, f'={get_column_letter(col)}{hire_total}', money)
for c in range(1, 8): ws.cell(r, c).fill = green
r += 1

sc(ws, r, 1, 'TOTAL REVENUE', b=True)
for col in range(3, 8): sc(ws, r, col, f'={get_column_letter(col)}{r-2}+{get_column_letter(col)}{r-1}', money)
navy_row(ws, r, 7)
proj_total = r
r += 1

sc(ws, r, 1, 'Goal', b=True)
for col in range(3, 8): sc(ws, r, col, '=$B$3', money)
r += 1

sc(ws, r, 1, 'OVER / (SHORT)', b=True)
ws.cell(r, 1).font = red_font
for col in range(3, 8):
    sc(ws, r, col, f'={get_column_letter(col)}{proj_total}-{get_column_letter(col)}{r-1}', money).font = Font(name='Arial', bold=True, size=14, color='FF0000')

set_widths(ws, [28, 12, 16, 16, 16, 16, 16])

# ============================================================
# SHEET 2: What to Expect (Years 1-3 with backup data)
# ============================================================
ws2 = wb.create_sheet('What to Expect')
ws2.sheet_properties.tabColor = '70AD47'

ws2['A1'] = 'What to Expect from a New Rep: Years 1-3'
ws2['A1'].font = Font(name='Arial', bold=True, size=18, color='143B87')
ws2.merge_cells('A1:H1')
ws2['A2'] = 'Based on 25 years of TSI invoice data | 200K+ invoices | 120+ reps | MEDIAN values used (see below)'
ws2['A2'].font = Font(name='Arial', italic=True, size=11, color='666666')
ws2.merge_cells('A2:H2')

# Why Median
r = 4
ws2.cell(r, 1, 'WHY MEDIAN INSTEAD OF AVERAGE?').font = Font(name='Arial', bold=True, size=12, color='FF0000')
ws2.merge_cells(f'A{r}:H{r}')
r += 1
why_median = [
    'The median is the "middle rep" — half did better, half did worse. It removes the effect of outliers.',
    'A few star reps (Kenney at $22M career, Hightower at $19M) pull the average up significantly.',
    'The median gives Denis a realistic expectation for a typical new hire, not the best-case scenario.',
    'Example: Inside Rep Year 1 — Average is $203K but Median is $146K. The $203K is inflated by a few reps who inherited huge books.',
    'Use median for budgeting. If the new hire beats it, that\'s upside. The Historical Data tab shows both for reference.',
]
for text in why_median:
    ws2.cell(r, 1, f'  {text}').font = Font(name='Arial', size=10)
    ws2.merge_cells(f'A{r}:H{r}')
    r += 1

r += 1
# Outside Reps section
ws2.cell(r, 1, 'OUTSIDE TERRITORY REP').font = Font(name='Arial', bold=True, size=14, color='143B87')
ws2.merge_cells(f'A{r}:H{r}')
r += 1
ws2.cell(r, 1, 'Sample: 8 confirmed outside reps (Lindsey Davis, Courtney Blue, Eric Schwarzel, Danielle Penge, Michael Woessner, Kara Klund, Melissa Fox, Bernie DeLacy)').font = Font(name='Arial', italic=True, size=9, color='666666')
ws2.merge_cells(f'A{r}:H{r}')
r += 1

for c, h in enumerate(['Metric', 'Year 1', 'Year 2', 'Year 3', 'How We Got This'], 1):
    ws2.cell(r, c, h)
hdr(ws2, r, 5)
r += 1

outside_metrics = [
    ('Total Revenue', '$357K', '$400K', '$210K', 'Median of 8 outside reps — removes Lindsey Davis outlier'),
    ('  FFS Revenue', '$357K', '$400K', '$210K', '100% of outside rep revenue was FFS'),
    ('  Contract Revenue', '$0', '$0', '$0', 'Outside reps landed zero contracts in first 5 years'),
    ('Net New Revenue', '$16K', '$15K', '$55K', 'Self-sourced business — median (net new slower than avg)'),
    ('Net New % of Total', '5%', '4%', '26%', 'Net new doesn\'t meaningfully kick in until Year 3'),
    ('Inherited Revenue', '$341K', '$385K', '$155K', 'Book they were given — carries them early'),
    ('Unique Clients', '20', '17', '14', 'Fewer clients but bigger accounts'),
    ('Net New Clients', '4', '4', '4', 'Slow client acquisition'),
    ('Avg Revenue/Client', '$18K', '$24K', '$15K', 'Big ticket — hospital systems, medical centers'),
    ('Instrument Mix', '45-78% Flex', '', '', 'Flexible endoscopes dominate, rigid 10-30%'),
]

for i, (metric, y1, y2, y3, source) in enumerate(outside_metrics):
    sc(ws2, r, 1, metric, b=not metric.startswith(' '))
    sc(ws2, r, 2, y1)
    sc(ws2, r, 3, y2)
    sc(ws2, r, 4, y3)
    sc(ws2, r, 5, source).font = Font(name='Arial', italic=True, size=9, color='666666')
    if i % 2 == 0:
        for c in range(1, 6): ws2.cell(r, c).fill = green
    r += 1

r += 1
ws2.cell(r, 1, 'The outside rep story: They coast on an inherited book for 2 years. Year 3 is when their own hunting pays off. By Year 4-5 they\'re $800K-$1.5M. But they need a territory with existing clients to start.').font = Font(name='Arial', italic=True, size=10, color='333333')
ws2.merge_cells(f'A{r}:H{r}')

# Inside Reps section
r += 2
ws2.cell(r, 1, 'INSIDE / PHONE REP').font = Font(name='Arial', bold=True, size=14, color='143B87')
ws2.merge_cells(f'A{r}:H{r}')
r += 1
ws2.cell(r, 1, 'Sample: 100+ inside reps over 25 years. Includes current top producers like Kenney, Hightower, Glavin, Cook.').font = Font(name='Arial', italic=True, size=9, color='666666')
ws2.merge_cells(f'A{r}:H{r}')
r += 1

for c, h in enumerate(['Metric', 'Year 1', 'Year 2', 'Year 3', 'How We Got This'], 1):
    ws2.cell(r, c, h)
hdr(ws2, r, 5)
r += 1

inside_metrics = [
    ('Total Revenue', '$146K', '$147K', '$222K', 'Median of 101 inside reps — removes outliers like Kenney/Hightower'),
    ('  FFS Revenue', '$140K', '$142K', '$210K', 'FFS dominates early'),
    ('  Contract Revenue', '$6K', '$5K', '$12K', 'Small early; grows with relationship depth'),
    ('Net New Revenue', '$32K', '$48K', '$89K', 'Self-sourced business (median)'),
    ('Net New % of Total', '22%', '33%', '40%', 'Higher net new % than outside from day 1'),
    ('Inherited Revenue', '$114K', '$99K', '$133K', 'Inherited book they start with'),
    ('Unique Clients', '19', '18', '22', 'More clients, smaller accounts'),
    ('Net New Clients', '8', '9', '11', 'Faster client acquisition than outside'),
    ('Avg Revenue/Client', '$8K', '$8K', '$10K', 'Lower per-client but more consistent'),
    ('Contract Potential', '0 contracts', '~1', '~2', 'Inside reps land contracts through relationships'),
    ('Retention Risk', '', '', '', '53% leave by Year 3. Only 19% make it to Year 5.'),
]

for i, (metric, y1, y2, y3, source) in enumerate(inside_metrics):
    sc(ws2, r, 1, metric, b=not metric.startswith(' '))
    sc(ws2, r, 2, y1)
    sc(ws2, r, 3, y2)
    sc(ws2, r, 4, y3)
    sc(ws2, r, 5, source).font = Font(name='Arial', italic=True, size=9, color='666666')
    if i % 2 == 0:
        for c in range(1, 6): ws2.cell(r, c).fill = light
    r += 1

r += 1
ws2.cell(r, 1, 'The inside rep story: Lower ceiling but faster to produce. They land contracts by Year 2 that outside reps never do. The big risk is turnover — more than half leave before Year 3. The ones who stay become your Kenneys and Hightowers.').font = Font(name='Arial', italic=True, size=10, color='333333')
ws2.merge_cells(f'A{r}:H{r}')

# Head to head comparison
r += 3
ws2.cell(r, 1, 'HEAD TO HEAD: YEAR 1-3').font = Font(name='Arial', bold=True, size=14, color='FF0000')
ws2.merge_cells(f'A{r}:H{r}')
r += 1
for c, h in enumerate(['', 'Outside Rep', 'Inside Rep', 'Winner'], 1):
    ws2.cell(r, c, h)
hdr(ws2, r, 4)
r += 1

h2h = [
    ('Year 1 Revenue', '$357K', '$146K', 'Outside (2.4x)'),
    ('Year 3 Revenue', '$210K', '$222K', 'Even — inside slightly edges'),
    ('Year 1 Net New', '$24K', '$32K', 'Inside'),
    ('Year 3 Net New', '$55K', '$89K', 'Inside (1.6x)'),
    ('Contract Revenue (Yr3)', '$0', '$12K', 'Inside'),
    ('Client Acquisition Speed', '5-6/yr', '9-12/yr', 'Inside'),
    ('Revenue per Client', '$15-33K', '$9-13K', 'Outside'),
    ('Retention (5yr)', 'N/A (small sample)', '19%', 'Outside (anecdotally better)'),
    ('Time to $500K', '~Year 3', '~Year 5+', 'Outside'),
    ('Upfront Book Needed?', 'Yes (critical)', 'Nice to have', 'Inside (lower risk)'),
]
for i, (metric, out, ins, winner) in enumerate(h2h):
    sc(ws2, r, 1, metric, b=True)
    sc(ws2, r, 2, out)
    sc(ws2, r, 3, ins)
    sc(ws2, r, 4, winner, b=True)
    if 'Outside' in winner:
        ws2.cell(r, 4).fill = green
    elif 'Inside' in winner:
        ws2.cell(r, 4).fill = light
    if i % 2 == 0:
        for c in range(1, 4): ws2.cell(r, c).fill = alt_fill
    r += 1

# Data sources
r += 2
ws2.cell(r, 1, 'DATA SOURCES').font = section
r += 1
sources = [
    'tblInvoice: 200K+ invoices from 1994-2026. FFS revenue attributed to rep on invoice at time of billing.',
    'tblContractDepartmentInvoiceSchedule: 7,066 contract billing rows, $35.4M total. Attributed to contract\'s own sales rep.',
    'Outside reps identified by Joe Brassell from institutional knowledge — not a system flag.',
    'Net new = client\'s first-ever invoice was under this rep. Inherited = client existed before this rep.',
    'Revenue averaged 2024-2025 for current roster to smooth contract spikes (e.g. Surgical Solutions).',
    '8 outside territory reps confirmed: Davis, Schwarzel, Blue, Penge, Woessner, Klund, Fox, DeLacy.',
    '120 inside reps analyzed after removing ghost keys (<$300K over 5+ year span).',
]
for s in sources:
    sc(ws2, r, 1, f'  - {s}')
    ws2.merge_cells(f'A{r}:H{r}')
    r += 1

set_widths(ws2, [26, 14, 14, 14, 50])

# ============================================================
# SHEET 3: Inside vs Outside (compact reference)
# ============================================================
ws3 = wb.create_sheet('Historical Data')
ws3.sheet_properties.tabColor = '4472C4'
ws3['A1'] = 'Historical Reference: Full Year 1-10 Data'
ws3['A1'].font = Font(name='Arial', bold=True, size=16, color='143B87')
ws3.merge_cells('A1:G1')

r = 3
ws3.cell(r, 1, 'OUTSIDE REP REVENUE BY TENURE YEAR').font = section
r += 1
for c, h in enumerate(['', 'Yr1', 'Yr2', 'Yr3', 'Yr4', 'Yr5', '# Reps'], 1): ws3.cell(r, c, h)
hdr(ws3, r, 7)
r += 1
sc(ws3, r, 1, 'Total Revenue', b=True)
for c, v in enumerate([356904, 400282, 209953, 814695, 1455687], 2): sc(ws3, r, c, v, money)
sc(ws3, r, 7, '8/8/5/2/1 (median)')
for c in range(1, 8): ws3.cell(r, c).fill = green

r += 2
ws3.cell(r, 1, 'INSIDE REP REVENUE BY TENURE YEAR').font = section
r += 1
for c, h in enumerate(['', 'Yr1', 'Yr2', 'Yr3', 'Yr4', 'Yr5', 'Yr6', 'Yr7', 'Yr8', 'Yr9', 'Yr10'], 1): ws3.cell(r, c, h)
hdr(ws3, r, 11)
r += 1

inside_yrs = [
    ('Median (used in planner)', [145928, 146988, 222298, 273829, 364239, 537314, 500393, 567328, 974340, 517939]),
    ('Average', [202513, 231005, 311416, 349960, 448195, 576277, 591763, 581976, 769530, 535461]),
    ('  FFS', [193162, 224668, 295391, 322212, 426483, 563926, 584169, 557151, 717057, 481599]),
    ('  Contract', [9351, 6336, 16025, 27747, 21712, 12352, 7594, 24825, 52473, 53862]),
    ('# Reps', [101, 77, 44, 33, 22, 15, 14, 12, 7, 8]),
]
for label, vals in inside_yrs:
    sc(ws3, r, 1, label, b=not label.startswith(' '))
    fmt = money if label != '# Reps' else '#,##0'
    for c, v in enumerate(vals, 2): sc(ws3, r, c, v, fmt)
    if 'Total' in label:
        for c in range(1, 12): ws3.cell(r, c).fill = light
    r += 1

r += 2
ws3.cell(r, 1, 'RETENTION: INSIDE REP TENURE').font = section
r += 1
for c, h in enumerate(['Tenure', '# Reps', '% of Total', 'Cumulative % Gone'], 1): ws3.cell(r, c, h)
hdr(ws3, r, 4)
r += 1
for tenure, n, p, cum, fill in [
    ('1-2 years', 25, 0.21, 0.21, PatternFill('solid', fgColor='FFC7CE')),
    ('2-3 years', 38, 0.32, 0.53, orange),
    ('3-4 years', 22, 0.18, 0.71, PatternFill('solid', fgColor='FFEB9C')),
    ('4-5 years', 12, 0.10, 0.81, light),
    ('5+ years', 23, 0.19, 1.00, green),
]:
    sc(ws3, r, 1, tenure, b=True); sc(ws3, r, 2, n); sc(ws3, r, 3, p, pct); sc(ws3, r, 4, cum, pct)
    for c in range(1, 5): ws3.cell(r, c).fill = fill
    r += 1

r += 2
ws3.cell(r, 1, 'COMPANY REVENUE').font = section
r += 1
for c, h in enumerate(['Year', 'FFS', 'Contract', 'Total', 'Contract %'], 1): ws3.cell(r, c, h)
hdr(ws3, r, 5)
r += 1
for yr, ffs, con in [(2024, 8958775, 5323596), (2025, 11771287, 6613925)]:
    sc(ws3, r, 1, yr); ws3.cell(r, 1).number_format = '0'
    sc(ws3, r, 2, ffs, money); sc(ws3, r, 3, con, money)
    sc(ws3, r, 4, f'=B{r}+C{r}', money); sc(ws3, r, 5, f'=C{r}/D{r}', pct)
    r += 1

for c in range(1, 12): ws3.column_dimensions[get_column_letter(c)].width = 14
ws3.column_dimensions['A'].width = 20

# Save
out = r'C:\tmp\tsi-sales-rep-planner-v8.xlsx'
wb.save(out)
print(f'Saved to {out}')
print('Sheets:', [s.title for s in wb.worksheets])
